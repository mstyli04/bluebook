"""SheetWriter: the single path every sheet in the workbook is written through.

It wraps a worksheet and a :class:`~bluebook.workbook.layout.Layout` and
exposes a handful of row methods (`title`, `year_header`, `input_row`,
`formula_row`, `blank`). Each call writes at the writer's internal cursor and
advances it, so callers never track row numbers themselves; `input_row` and
`formula_row` additionally register their key in the `Layout` at the row they
wrote, so later sheets can address the value with `layout.ref(...)` instead
of a hardcoded cell reference.

A `SheetWriter` is single-mode: constructed with `historical=True` it writes
row values into `HIST_COLS` (three historical years), constructed with
`historical=False` (the default) it writes into `FCST_COLS` (five forecast
years). `year_header` alone can override the mode per call (its own
`historical` argument), which lets one writer still label a historical block
of columns it does not itself write values into.

--------------------------------------------------------------------------
Explicit column sets (`cols`)
--------------------------------------------------------------------------
Two sheets Task 13 adds are not laid out by year at all, so neither column
block fits them: **Comps** puts one COMPANY per column (the five peers, then
three statistics columns beside them) and **Sensitivity** puts one axis value
per column across a 5x5 grid whose position the plan's test fixes at
``D5:H9``. Both therefore need columns the two year blocks do not name.

`cols` is how they get them, and it is deliberately additive rather than a
change to anything Task 12 fixed: the constructor takes an optional default
and `year_header` / `input_row` / `formula_row` take an optional per-call
override, and with neither supplied every writer behaves exactly as before.
What it does NOT do is open a side door around the guards below — a row
written into explicit columns still goes through `_write_row`, so it is still
registered in the `Layout` at the row the cursor was on, still styled by the
same fonts, and still refused by the hardcode rule on a sheet that has no
licence to hold constants. Writing those cells straight onto the worksheet
instead (`writer.ws["D5"] = ...`) would have skipped all three, which is the
bypass `sheet_statements.py`'s docstring warns against.

Row 1 is always the sheet title, row 2 the year header (the plan's Global
Constraints state this explicitly) — `title()` and `year_header()` each
advance the cursor by exactly one row, so the first row a caller registers
through `input_row`/`formula_row` lands on row 3. Sensitivity is the one
sheet with no year axis to put there: it calls `blank()` for row 2 rather
than writing a header that would label its axis columns as years.

`formula_row` and `input_row` enforce the hardcode rule from `styles.py`'s
`HARDCODE_ALLOWED` at the point of writing, not just at a later scan: on any
sheet outside that set, `formula_row` rejects any value that is not a
`"="`-prefixed formula string, and `input_row` is rejected outright (its
purpose — writing constants — is only legal on the allowed sheets).
"""

from __future__ import annotations

from typing import Sequence

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from bluebook.workbook.layout import FCST_COLS, HIST_COLS, Layout
from bluebook.workbook.styles import (
    FORMULA_FONT,
    HARDCODE_ALLOWED,
    HEADER_FONT,
    INPUT_FONT,
    LINK_FONT,
    MONEY_FORMAT,
    TITLE_FONT,
)

LABEL_COL = "B"


class SheetWriter:
    """Writes one worksheet top to bottom, registering rows in a Layout."""

    def __init__(
        self,
        worksheet: Worksheet,
        layout: Layout,
        *,
        historical: bool = False,
        cols: Sequence[str] | None = None,
    ) -> None:
        self.ws = worksheet
        self.layout = layout
        self.sheet = worksheet.title
        self.historical = historical
        self.cols = tuple(cols) if cols is not None else None
        self.row = 1

    def columns(self, cols: Sequence[str] | None = None) -> tuple[str, ...]:
        """The columns a row's values are written into, most specific first.

        A per-call `cols` wins over the writer's own default, which wins over
        the year block the writer's mode selects. One resolution order, used by
        both `year_header` and `_write_row`, so a header cannot end up over a
        different set of columns from the rows beneath it.
        """
        if cols is not None:
            return tuple(cols)
        if self.cols is not None:
            return self.cols
        return HIST_COLS if self.historical else FCST_COLS

    def title(self, text: str) -> None:
        """Write the sheet title in A1. The year header follows immediately on row 2."""
        cell = self.ws[f"A{self.row}"]
        cell.value = text
        cell.font = TITLE_FONT
        self.row += 1

    def year_header(
        self,
        labels: Sequence[str],
        historical: bool | None = None,
        *,
        cols: Sequence[str] | None = None,
    ) -> int:
        """Write column labels across the sheet's data columns, bold.

        Returns the row written, so a caller that needs to address the header
        later reads the position rather than recomputing it.
        """
        columns = (
            self.columns(cols)
            if historical is None
            else (HIST_COLS if historical else FCST_COLS)
        )
        row = self.row
        for index, label in enumerate(labels):
            cell = self.ws[f"{columns[index]}{row}"]
            cell.value = label
            cell.font = HEADER_FONT
        self.row += 1
        return row

    def input_row(
        self,
        key: str,
        label: str,
        values: Sequence,
        fmt: str = MONEY_FORMAT,
        *,
        cols: Sequence[str] | None = None,
    ) -> int:
        """Write a row of hardcoded values in blue and register it in the Layout.

        Raises ValueError if this sheet is not in `HARDCODE_ALLOWED` — writing
        constants at all is only legal on the five sheets that list permits.
        """
        if self.sheet not in HARDCODE_ALLOWED:
            raise ValueError(
                f"sheet {self.sheet!r} is not in HARDCODE_ALLOWED "
                f"{sorted(HARDCODE_ALLOWED)!r}: input_row({key!r}, ...) may not "
                f"write hardcoded constants here"
            )
        return self._write_row(key, label, values, fmt, INPUT_FONT, cols)

    def formula_row(
        self,
        key: str,
        label: str,
        formulas: Sequence[str],
        fmt: str = MONEY_FORMAT,
        *,
        is_link: bool = False,
        cols: Sequence[str] | None = None,
    ) -> int:
        """Write a row of formulas (black, or green if `is_link`) and register it.

        On any sheet not in `HARDCODE_ALLOWED`, every value must be a
        `"="`-prefixed formula string; a bare constant raises ValueError
        naming the sheet, the key and the offending value, so it is caught
        here rather than by a later scan of the finished workbook.
        """
        if self.sheet not in HARDCODE_ALLOWED:
            for value in formulas:
                if not (isinstance(value, str) and value.startswith("=")):
                    raise ValueError(
                        f"sheet {self.sheet!r}, row {key!r}: formula_row may only "
                        f"write formula strings here, got {value!r}"
                    )
        font = LINK_FONT if is_link else FORMULA_FONT
        return self._write_row(key, label, formulas, fmt, font, cols)

    def blank(self, rows: int = 1) -> None:
        """Advance the cursor without writing anything, leaving `rows` blank."""
        self.row += rows

    def _write_row(
        self,
        key: str,
        label: str,
        values: Sequence,
        fmt: str,
        font: Font,
        cols: Sequence[str] | None = None,
    ) -> int:
        row = self.row
        self.layout.register(self.sheet, key, row)
        self.ws[f"{LABEL_COL}{row}"] = label
        columns = self.columns(cols)
        for index, value in enumerate(values):
            cell = self.ws[f"{columns[index]}{row}"]
            cell.value = value
            cell.font = font
            cell.number_format = fmt
        self.row += 1
        return row
