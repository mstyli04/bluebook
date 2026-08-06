"""What headless LibreOffice's iterative calculation will and will not resolve.

Task 2's spike established that LibreOffice honours the ``wb.calculation``
iteration settings openpyxl writes, and `INTEREST_BASIS = "average"` in
`schedules/debt.py` was chosen on that basis. The spike used one circular cell
pair. This file measures the two shapes the real workbook actually needs, and
LibreOffice resolves neither:

* a **branched** circular group — one cell inside the loop read by two cells
  that are also inside it — converges with the second branch frozen at its
  seed value, giving a stable but wrong answer;
* a **chain** of circular groups — which a multi-year debt schedule is, one
  group per year — resolves the first group correctly and then leaves the rest
  reading frozen inputs.

Both are pinned here rather than only described in prose, because the
conclusion they support is load-bearing: a five-year average-balance debt
schedule cannot be made to recalculate correctly in LibreOffice by rearranging
its rows, so the interest basis is an owner decision rather than a layout
problem. See `workbook/sheet_schedules.py`.

If a future LibreOffice fixes either behaviour, the assertions below will fail
with a message saying so; that is a prompt to revisit `INTEREST_BASIS`, not a
regression.
"""

from pathlib import Path

import openpyxl
import pytest

from bluebook.recalc import recalc_values

RATE = 0.055
OPENING = 25.0
CASH = 100.0
ITERATE_COUNT = 100
ITERATE_DELTA = 0.0001
# `iterateDelta` is the convergence threshold LibreOffice stops at, so a group
# it resolves correctly still lands a little short of the exact fixed point
# (~2e-6 as measured). Comparisons against the analytical answer are made to
# this tolerance, not to machine precision.
CONVERGED_TOLERANCE = 1e-4


def _iterating_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.calculation.iterate = True
    wb.calculation.iterateCount = ITERATE_COUNT
    wb.calculation.iterateDelta = ITERATE_DELTA
    return wb


def _solve_year(opening: float) -> tuple[float, float]:
    """The fixed point of one year: interest on average debt, cash repays it."""
    closing = opening
    for _ in range(1000):
        interest = (opening + closing) / 2 * RATE
        closing = opening + CASH - interest
    return interest, closing


def test_a_single_circular_group_resolves_correctly(tmp_path: Path):
    """The shape Task 2's spike verified, restated in this file's terms."""
    wb = _iterating_workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = RATE
    ws["A2"] = CASH
    ws["F5"] = OPENING
    ws["F6"] = "=AVERAGE(S!F5,S!F8)*S!$A$1"   # interest, circular
    ws["F8"] = "=S!F5+S!$A$2-S!F6"            # closing, circular
    path = tmp_path / "single.xlsx"
    wb.save(path)

    values = recalc_values(path)["S"]
    interest, closing = _solve_year(OPENING)
    assert values["F6"] == pytest.approx(interest, abs=CONVERGED_TOLERANCE)
    assert values["F8"] == pytest.approx(closing, abs=CONVERGED_TOLERANCE)


def test_a_branched_circular_group_freezes_its_second_branch(tmp_path: Path):
    """`X` feeds two cells that both feed `Y`; only one of them is iterated.

    The correct answer solves X = (A + Y) / 2 * r with Y = A + K - P - Q and
    P = Q = X. LibreOffice returns a self-consistent answer for every cell
    except Q, which holds the value X took on the seeding pass.
    """
    wb = _iterating_workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = RATE
    ws["A2"] = CASH
    ws["F5"] = OPENING
    ws["F6"] = "=AVERAGE(S!F5,S!F9)*S!$A$1"   # X, circular
    ws["F7"] = "=S!F6"                        # P, first branch
    ws["F8"] = "=S!F6"                        # Q, second branch
    ws["F9"] = "=S!F5+S!$A$2-S!F7-S!F8"       # Y, circular
    path = tmp_path / "branched.xlsx"
    wb.save(path)

    values = recalc_values(path)["S"]
    # P and Q are the same formula. That they disagree is the whole finding:
    # one of them is being iterated and the other is not. (As measured, Q holds
    # 4.125 — X evaluated on the pass where P and Q were still nil, so Y read
    # A + K = 125 — while P holds the iterated value.)
    assert values["F7"] != pytest.approx(values["F8"], abs=CONVERGED_TOLERANCE), (
        "P and Q are identical formulas and now agree, so LibreOffice is "
        "iterating both branches — revisit INTEREST_BASIS and the workbook's "
        "debt schedule."
    )
    # And the answer is wrong, not merely unfinished: X should satisfy
    # X = (A + (A + K - 2X)) / 2 * r.
    true_x = (2 * OPENING + CASH) * RATE / 2 / (1 + RATE)
    assert abs(values["F6"] - true_x) > 1e-3, (
        f"LibreOffice now resolves a branched circular group (X={values['F6']}, "
        f"true {true_x}) — revisit INTEREST_BASIS and the workbook's debt schedule."
    )


def test_a_chain_of_circular_groups_resolves_only_the_first(tmp_path: Path):
    """Five years, each its own circular group, chained by the closing balance.

    This is the shape of a multi-year average-balance debt schedule. FY1 comes
    out exact; from FY2 on, the interest cell is left behind and the closing
    balances drift further from the true fixed point every year.
    """
    wb = _iterating_workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = RATE
    ws["A2"] = CASH
    for index, col in enumerate("FGHIJ"):
        prior = "FGHIJ"[index - 1]
        ws[f"{col}5"] = OPENING if index == 0 else f"=S!{prior}8"
        ws[f"{col}6"] = f"=AVERAGE(S!{col}5,S!{col}8)*S!$A$1"
        ws[f"{col}8"] = f"=S!{col}5+S!$A$2-S!{col}6"
    path = tmp_path / "chained.xlsx"
    wb.save(path)

    values = recalc_values(path)["S"]
    expected = []
    opening = OPENING
    for _ in range(5):
        _, closing = _solve_year(opening)
        expected.append(closing)
        opening = closing

    assert values["F8"] == pytest.approx(expected[0], abs=CONVERGED_TOLERANCE), (
        "the first circular group in a chain should still resolve"
    )
    later_error = max(abs(values[f"{col}8"] - expected[i]) for i, col in enumerate("GHIJ", 1))
    assert later_error > 1e-3, (
        f"LibreOffice now resolves a chain of circular groups (worst error "
        f"{later_error:.3g}) — revisit INTEREST_BASIS and the workbook's debt "
        f"schedule, which was laid out on the assumption that it does not."
    )
