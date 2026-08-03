"""Spike: does LibreOffice honour iterative calculation from openpyxl?

Models one year of circular interest:
    interest = rate * average(opening debt, closing debt)
    closing debt = opening debt - (cash_before_interest - interest)

With opening=100, rate=10%, cash_before_interest=30, the converged solution is
closing = 100 - 30 + 0.10 * (100 + closing) / 2, i.e. closing ≈ 78.95,
interest ≈ 8.95.
"""

from pathlib import Path

import openpyxl
import pytest

from bluebook.recalc import recalc_values


def test_libreoffice_resolves_circular_reference(tmp_path: Path):
    wb = openpyxl.Workbook()
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100
    wb.calculation.iterateDelta = 0.0001

    ws = wb.active
    ws.title = "Debt"
    ws["B1"] = 100.0   # opening debt
    ws["B2"] = 0.10    # interest rate
    ws["B3"] = 30.0    # cash before interest
    ws["B4"] = "=B2*AVERAGE(B1,B5)"        # interest (circular)
    ws["B5"] = "=B1-(B3-B4)"               # closing debt (circular)

    path = tmp_path / "circular.xlsx"
    wb.save(path)

    values = recalc_values(path)["Debt"]
    assert values["B5"] == pytest.approx(78.95, abs=0.05)
    assert values["B4"] == pytest.approx(8.95, abs=0.05)
