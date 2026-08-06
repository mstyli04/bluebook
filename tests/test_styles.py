import openpyxl
import pytest

from bluebook.workbook.layout import Layout
from bluebook.workbook.sheet import SheetWriter
from bluebook.workbook.styles import (
    FORMULA_FONT,
    HARDCODE_ALLOWED,
    INPUT_FONT,
    LINK_FONT,
)


def test_fonts_use_the_agreed_colours():
    assert INPUT_FONT.color.rgb == "FF0000FF"
    assert FORMULA_FONT.color.rgb == "FF000000"
    assert LINK_FONT.color.rgb == "FF008000"


def test_input_row_writes_values_in_blue_and_registers_its_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)
    writer.title("Assumptions")
    writer.year_header(["FY2026", "FY2027", "FY2028", "FY2029", "FY2030"])
    writer.input_row("tax_rate", "Tax rate", [0.25] * 5)

    row = layout.row_of("Assumptions", "tax_rate")
    assert ws[f"F{row}"].value == pytest.approx(0.25)
    assert ws[f"F{row}"].font.color.rgb == "FF0000FF"


def test_formula_row_writes_formulas_in_black():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)
    writer.title("Income Statement")
    writer.year_header(["FY2026", "FY2027", "FY2028", "FY2029", "FY2030"])
    writer.formula_row("revenue", "Revenue", ["=1+1"] * 5)

    row = layout.row_of("IS", "revenue")
    assert ws[f"F{row}"].value == "=1+1"
    assert ws[f"F{row}"].font.color.rgb == "FF000000"


def test_title_and_year_header_land_on_the_fixed_rows_the_plan_requires():
    # Global Constraints: "Row 1 is the sheet title, row 2 the year header."
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)
    writer.title("Assumptions")
    writer.year_header(["FY2026", "FY2027", "FY2028", "FY2029", "FY2030"])
    writer.input_row("tax_rate", "Tax rate", [0.25] * 5)

    assert ws["A1"].value == "Assumptions"
    assert ws["F2"].value == "FY2026"
    assert layout.row_of("Assumptions", "tax_rate") == 3


def test_hardcode_allowed_matches_the_plans_five_sheets():
    assert HARDCODE_ALLOWED == {
        "Assumptions", "Historicals", "Comps", "Cover", "Checks",
    }


def test_input_row_rejects_sheets_outside_the_hardcode_allow_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)

    with pytest.raises(ValueError, match="IS") as exc:
        writer.input_row("revenue", "Revenue", [123.4] * 5)
    assert "revenue" in str(exc.value)


def test_formula_row_rejects_bare_constants_outside_the_hardcode_allow_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)

    with pytest.raises(ValueError) as exc:
        writer.formula_row("revenue", "Revenue", [123.4, 234.5, 111.1, 90.0, 88.8])
    message = str(exc.value)
    assert "IS" in message
    assert "revenue" in message
    assert "123.4" in message


def test_formula_row_allows_formula_strings_outside_the_hardcode_allow_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=False)

    # Should not raise: every value is a formula string.
    writer.formula_row("revenue", "Revenue", ["=1+1"] * 5)


def test_input_row_and_formula_row_are_unrestricted_on_allow_listed_sheets():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historicals"
    layout = Layout()
    writer = SheetWriter(ws, layout, historical=True)

    # Should not raise on an allow-listed sheet.
    writer.input_row("revenue", "Revenue", [1508.3, 1802.9, 2013.1])
