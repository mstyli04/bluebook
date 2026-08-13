"""The Checks sheet: that it reads TRUE, and that it could read FALSE.

This file was in the plan from the start and was never written. Its absence
was the largest hole in the suite: the eight checks sit on the second tab and
are the first thing a reader sees, and **nothing failed when one of them read
FALSE**. Two independent reviews found it the same way — set
`TERMINAL_SHARE_CEILING` to 0.50, regenerate, and 298 tests pass while the
delivered workbook shows a FALSE on tab two.

The error scan in `test_workbook_sheets.py` does not cover this: `FALSE` is a
correct evaluation of a formula, not an error value. Nor did the cross-check,
which excluded `Checks` — on the strength of a comment claiming this file
already existed.

Two properties are worth separating:

* every check reads TRUE, in every scenario, on the workbook as shipped; and
* the sheet is capable of reading FALSE at all.

The second matters because a sheet of checks that cannot fail is worse than no
sheet — it is a claim of verification that has not been made. This project has
already shipped one: `check_cash_ties` compared a cell to itself until Task
14's review caught it.
"""

from pathlib import Path

import openpyxl
import pytest

from bluebook.assumptions import SCENARIOS
from bluebook.inputs.greggs import GREGGS_HISTORICALS
from bluebook.recalc import recalc_values
from bluebook.reference import build_model
from bluebook.workbook.build import build_workbook

# Literal addresses, per this suite's convention: a row resolved through the
# layout would follow the writer wherever it put the row, so it could not
# catch a check that moved out from under its own label.
FIRST_CHECK_ROW = 3
LAST_CHECK_ROW = 10
RESULT_COL = "D"

# The label each check row must carry, in order. Written out rather than read
# from the sheet — this is the list of things the workbook claims to verify,
# and a check silently disappearing is exactly what this pins.
CHECK_LABELS = (
    "Balance sheet balances in every forecast year",
    "Closing cash agrees across all three constructions of it, every year",
    "Closing borrowings never negative (no accidental cash-as-debt)",
    "Cash never falls below the £50m minimum the revolver defends",
    "Perpetuity growth is below WACC (the Gordon formula is defined)",
    "Terminal value is below 97% of enterprise value",
    "No forecast revenue growth outside +/-50%, including the step off the "
    "last reported year",
    "Peak borrowings are an assumed facility upsize, not a solvency problem: "
    "lease-inclusive net debt / EBITDA below 2.0x every year",
)

# The three figures behind the financing disclosure, which the Cover note and
# the README both quote.
PEAK_BORROWINGS_CELL = "C13"
PEAK_LEVERAGE_CELL = "C14"
PEAK_GROSS_LEVERAGE_CELL = "C15"


@pytest.fixture(scope="module")
def base_workbook_path(tmp_path_factory) -> Path:
    return build_workbook(
        GREGGS_HISTORICALS, "Base", tmp_path_factory.mktemp("checks") / "model.xlsx"
    )


def _check_results(path: Path) -> list:
    values = recalc_values(path)["Checks"]
    return [values.get(f"{RESULT_COL}{row}")
            for row in range(FIRST_CHECK_ROW, LAST_CHECK_ROW + 1)]


def test_the_sheet_carries_exactly_the_checks_it_claims_to(base_workbook_path):
    ws = openpyxl.load_workbook(base_workbook_path)["Checks"]
    labels = [ws[f"B{row}"].value
              for row in range(FIRST_CHECK_ROW, LAST_CHECK_ROW + 1)]
    assert labels == list(CHECK_LABELS)
    # ...and nothing follows them in the result column, so a ninth check
    # cannot be added without this list noticing.
    assert ws[f"{RESULT_COL}{LAST_CHECK_ROW + 1}"].value is None


@pytest.mark.parametrize("scenario", ["Bear", "Base", "Bull"])
def test_every_check_reads_true(tmp_path, scenario):
    """The claim the workbook makes on its second tab, in every scenario."""
    path = build_workbook(GREGGS_HISTORICALS, scenario, tmp_path / f"{scenario}.xlsx")
    results = _check_results(path)
    failures = [label for label, result in zip(CHECK_LABELS, results)
                if result is not True]
    assert not failures, (
        f"{scenario}: {len(failures)} check(s) do not read TRUE:\n  "
        + "\n  ".join(failures)
        + f"\n(raw results: {results})"
    )


def test_a_broken_model_turns_the_checks_false(base_workbook_path, tmp_path):
    """The sheet must be capable of the answer it exists to give.

    Without this, `test_every_check_reads_true` is a test whose only observed
    behaviour is passing, which is the shape of every vacuous check this
    project has found. The mutation is deliberately crude — break the balance
    sheet — because the point is that the sheet reacts at all, not which rows
    react.
    """
    wb = openpyxl.load_workbook(base_workbook_path)
    ws = wb["BS"]
    # Inflate one asset. The balance check, and the ties that depend on it,
    # must notice.
    ws["F3"] = f"=({ws['F3'].value[1:]})+100"
    broken = tmp_path / "broken.xlsx"
    wb.save(broken)

    results = _check_results(broken)
    assert False in results, (
        f"breaking the balance sheet changed no check: {results}"
    )


def test_the_disclosure_figures_match_the_model(base_workbook_path):
    """The three numbers the Cover note and the README quote.

    `Checks!C13:C15` are the evidence for the RCF-upsize disclosure. They were
    written to the sheet and checked by nothing, while the prose quoting them
    was checked by nobody either — the combination that produced the leverage
    basis error Task 16 had to correct.
    """
    values = recalc_values(base_workbook_path)["Checks"]
    model = build_model(GREGGS_HISTORICALS, SCENARIOS["Base"])

    borrowings = model.balance_sheet["borrowings"]
    cash = model.balance_sheet["cash"]
    leases = model.balance_sheet["lease_liabilities"]
    ebitda = model.ebitda

    assert values[PEAK_BORROWINGS_CELL] == pytest.approx(max(borrowings), abs=1e-9)
    assert values[PEAK_LEVERAGE_CELL] == pytest.approx(
        max((b - c + l) / e for b, c, l, e in zip(borrowings, cash, leases, ebitda)),
        abs=1e-9,
    )
    assert values[PEAK_GROSS_LEVERAGE_CELL] == pytest.approx(
        max(b / e for b, e in zip(borrowings, ebitda)), abs=1e-9
    )
    # The disclosure's whole point: the two bases genuinely differ, and the
    # one on the sheet's check is the higher — the less flattering — of them.
    assert values[PEAK_LEVERAGE_CELL] > values[PEAK_GROSS_LEVERAGE_CELL]
