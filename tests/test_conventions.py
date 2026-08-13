"""House rules the whole workbook must obey, checked on the generated file.

Three of these four restate a guard that already exists elsewhere, on purpose.
`SheetWriter` refuses a constant on a sheet outside `HARDCODE_ALLOWED` at the
point of writing, and `test_workbook_sheets.py` checks titles and the absence
of hardcoded forecast cells while it is checking everything else. The value of
repeating them here is that this file is the one place a reader can look to see
what the conventions ARE, and that these versions inspect the finished
``.xlsx`` rather than the code that produced it — a bypass that wrote straight
onto the worksheet would satisfy the writer's guard and fail this one.

The fourth is not a restatement of anything. Nothing else in the suite would
fail if the workbook regained a circular reference, and that is the failure
mode the whole design is built to avoid: LibreOffice does not report a chained
or branched circularity, it returns a plausible, self-consistent, wrong number
(see `docs/superpowers/spike-circularity.md` and
`tests/test_libreoffice_iteration_limits.py`). Acyclicity was proved once, by a
reviewer parsing every formula into a reference graph, but that proof lived in
a review transcript. This puts it in the repository.
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string, get_column_letter

from bluebook.inputs.greggs import GREGGS_HISTORICALS
from bluebook.workbook.build import build_workbook
from bluebook.workbook.styles import HARDCODE_ALLOWED, INPUT_COLOR

# Column B holds row labels on every sheet; column A holds the title on row 1.
# Values start at column C, so that is where a convention scan starts.
FIRST_VALUE_COLUMN = 3


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory) -> Path:
    return build_workbook(
        GREGGS_HISTORICALS, "Base", tmp_path_factory.mktemp("conventions") / "model.xlsx"
    )


@pytest.fixture(scope="module")
def workbook(workbook_path):
    return openpyxl.load_workbook(workbook_path)


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_number(value) -> bool:
    """A hardcoded numeric constant.

    Both rules below are about NUMBERS, which is what `HARDCODE_ALLOWED`
    governs ("hardcoded numeric constants are permitted only on these five
    sheets") and what the blue/black convention signals. Text is neither: a
    column header on the football field, a source citation in Historicals'
    column L and a scenario name are all strings a reader can see are not
    figures, and sweeping them in would mean either exempting them one by one
    or repainting labels blue to satisfy a test.
    """
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _font_colour(cell) -> str | None:
    """The cell's font colour as an RGB string, or None if it has no RGB.

    openpyxl's `Color.rgb` is a typed descriptor that returns its own error
    text rather than a string when the colour is a theme or indexed reference,
    so the value has to be type-checked rather than trusted.
    """
    if not cell.font or not cell.font.color:
        return None
    rgb = cell.font.color.rgb
    return rgb if isinstance(rgb, str) else None


def _value_cells(ws):
    """Every written cell in the value region of `ws`, as (cell, value)."""
    for row in ws.iter_rows(min_row=2, min_col=FIRST_VALUE_COLUMN):
        for cell in row:
            if cell.value is not None:
                yield cell, cell.value


# --------------------------------------------------------------------------
# 1. Hardcoded numbers
# --------------------------------------------------------------------------


def test_only_the_allowed_sheets_hold_hardcoded_values(workbook):
    """Every value cell outside `HARDCODE_ALLOWED` is a formula.

    `HARDCODE_ALLOWED` is imported, never redefined here: two copies of the
    list is exactly how a sheet quietly acquires a licence it was never given.
    """
    offenders = [
        f"{ws.title}!{cell.coordinate} = {value!r}"
        for ws in workbook.worksheets
        if ws.title not in HARDCODE_ALLOWED
        for cell, value in _value_cells(ws)
        if _is_number(value)
    ]
    assert not offenders, (
        "hardcoded numbers on sheets that may not hold them "
        f"(allowed: {sorted(HARDCODE_ALLOWED)}):\n  " + "\n  ".join(offenders)
    )


# --------------------------------------------------------------------------
# 2. Input cells blue, formula cells not
# --------------------------------------------------------------------------


def test_input_cells_are_blue_and_formula_cells_are_not(workbook):
    """The colour convention a reader uses to tell an input from a calculation.

    Blue means "somebody typed this"; black and green mean "the workbook worked
    this out". A formula wearing input blue is the worse direction of the two
    errors — it invites a reader to overwrite a cell that drives the model.

    Row 2 is excluded because it is the year header, which is bold black on
    every sheet and is neither an input nor a calculation.
    """
    miscoloured: list[str] = []
    for ws in workbook.worksheets:
        for cell, value in _value_cells(ws):
            if cell.row == 2:
                continue
            colour = _font_colour(cell)
            if _is_formula(value) and colour == INPUT_COLOR:
                miscoloured.append(
                    f"{ws.title}!{cell.coordinate} is a formula in input blue"
                )
            elif _is_number(value) and colour != INPUT_COLOR:
                miscoloured.append(
                    f"{ws.title}!{cell.coordinate} is a hardcoded number but is "
                    f"not blue ({colour})"
                )
    assert not miscoloured, "colour convention broken:\n  " + "\n  ".join(miscoloured)


# --------------------------------------------------------------------------
# 3. Titles
# --------------------------------------------------------------------------


def test_every_sheet_has_a_non_empty_title_in_a1(workbook):
    for ws in workbook.worksheets:
        title = ws["A1"].value
        assert isinstance(title, str) and title.strip(), (
            f"{ws.title} has no title in A1 (found {title!r})"
        )


# --------------------------------------------------------------------------
# 4. Acyclicity
# --------------------------------------------------------------------------

# A cell reference, with or without a sheet prefix, with or without `$`, and
# either a single cell or a range. String literals are stripped before this is
# applied, so a scenario name inside `IF(...="Bull", ...)` cannot be mistaken
# for a reference.
#
# The negative lookbehind stops the tail of a longer token matching: without
# it, the `E31` in a hypothetical `SOMETHINGE31` would register as a reference
# to column E. `$` is explicitly permitted before a column letter because that
# is what an absolute reference looks like.
_REFERENCE = re.compile(
    r"""
    (?<![A-Za-z0-9_])
    (?:
        (?:'(?P<quoted>[^']+)'|(?P<bare>[A-Za-z_][A-Za-z0-9_.]*))!
    )?
    \$?(?P<col1>[A-Z]{1,3})\$?(?P<row1>[0-9]+)
    (?::\$?(?P<col2>[A-Z]{1,3})\$?(?P<row2>[0-9]+))?
    """,
    re.VERBOSE,
)
_STRING_LITERAL = re.compile(r'"[^"]*"')


def _references(formula: str, own_sheet: str) -> set[tuple[str, str]]:
    """Every (sheet, cell) a formula depends on, ranges expanded.

    Ranges are expanded rather than recorded as a unit because a cycle through
    one cell of a `MIN(...)` range is a cycle, and a range treated as a single
    node would either miss it or invent edges that do not exist.
    """
    text = _STRING_LITERAL.sub("", formula)
    found: set[tuple[str, str]] = set()
    for match in _REFERENCE.finditer(text):
        # A function name cannot be a reference: `SUM(` and friends never
        # match the column+row shape, but `LOG10(` would, so anything opening
        # a bracket is discarded.
        after = text[match.end() : match.end() + 1]
        if after == "(":
            continue
        sheet = match.group("quoted") or match.group("bare") or own_sheet
        first_col = column_index_from_string(match.group("col1"))
        first_row = int(match.group("row1"))
        if match.group("col2") is None:
            found.add((sheet, f"{match.group('col1')}{first_row}"))
            continue
        last_col = column_index_from_string(match.group("col2"))
        last_row = int(match.group("row2"))
        for col in range(min(first_col, last_col), max(first_col, last_col) + 1):
            for row in range(min(first_row, last_row), max(first_row, last_row) + 1):
                found.add((sheet, f"{get_column_letter(col)}{row}"))
    return found


def _reference_graph(wb) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """{(sheet, cell): the cells its formula reads}, for every formula."""
    graph: dict[tuple[str, str], set[tuple[str, str]]] = defaultdict(set)
    known = set(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not _is_formula(cell.value):
                    continue
                for sheet, address in _references(cell.value[1:], ws.title):
                    # A prefix that names no sheet in this workbook is not an
                    # edge — it is a broken reference, and the error scan in
                    # test_workbook_sheets.py is what reports that.
                    if sheet in known:
                        graph[(ws.title, cell.coordinate)].add((sheet, address))
    return graph


def _find_cycle(graph) -> list[tuple[str, str]] | None:
    """One cycle in `graph` as a node path, or None. Iterative: the graph is
    deep enough that recursion is a real stack-depth risk."""
    UNVISITED, IN_PROGRESS, DONE = 0, 1, 2
    state: dict[tuple[str, str], int] = defaultdict(int)

    for start in list(graph):
        if state[start] != UNVISITED:
            continue
        stack = [(start, iter(graph.get(start, ())))]
        path = [start]
        state[start] = IN_PROGRESS
        while stack:
            node, children = stack[-1]
            for child in children:
                if state[child] == IN_PROGRESS:
                    return path[path.index(child) :] + [child]
                if state[child] == UNVISITED:
                    state[child] = IN_PROGRESS
                    path.append(child)
                    stack.append((child, iter(graph.get(child, ()))))
                    break
            else:
                state[node] = DONE
                stack.pop()
                path.pop()
    return None


def test_the_workbook_has_no_circular_references(workbook):
    """The design's load-bearing property, enforced where the mistake happens.

    Interest is charged on OPENING debt balances precisely so this holds. If
    someone reinstates an average-balance interest row, the workbook still
    opens, still recalculates, and still prints a full set of numbers — the
    only thing that says otherwise is this test.
    """
    graph = _reference_graph(workbook)
    cycle = _find_cycle(graph)
    assert cycle is None, "circular reference:\n  " + "\n  -> ".join(
        f"{sheet}!{address}" for sheet, address in (cycle or [])
    )


def test_the_reference_graph_is_big_enough_to_be_meaningful(workbook):
    """Guards the acyclicity test against passing on an empty parse.

    A regex that silently stopped matching would make the test above pass on a
    graph with no edges at all, which is the "test that proves nothing" this
    project keeps finding.

    The workbook currently parses to 862 formula cells and 2,464 edges over
    1,084 nodes. (The 490 formulas and 1,050 edges quoted in the Task 12 ledger
    entry were the whole workbook as it stood then, before Tasks 13 and 14
    added the DCF, Sensitivity, Comps, LBO, Checks and Football Field sheets —
    the same measurement, not a different one.) The floors below are set well
    under that on purpose: they assert that parsing happened at all, and are
    not a count anybody has to maintain.
    """
    graph = _reference_graph(workbook)
    edges = sum(len(targets) for targets in graph.values())
    assert len(graph) > 300, f"only {len(graph)} formula cells parsed"
    assert edges > 700, f"only {edges} edges parsed"


def test_the_cycle_finder_reports_a_cycle_when_one_exists(workbook_path, tmp_path):
    """The detector, tested against a workbook that really is circular.

    Without this, `test_the_workbook_has_no_circular_references` is a test
    whose only observed behaviour is passing. Here the model is made circular
    the exact way the design forbids — interest charged on a balance that
    depends on interest — and the finder must say so.
    """
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Schedules"]
    interest_row = next(
        cell.row
        for cell in ws["B"]
        if isinstance(cell.value, str)
        and cell.value.lstrip().startswith("Interest on borrowings")
    )
    closing_row = next(
        cell.row
        for cell in ws["B"]
        if isinstance(cell.value, str)
        and cell.value.lstrip().startswith("Closing borrowings")
    )
    # Interest on the AVERAGE of opening and closing debt: the average-basis
    # row the spike rejected.
    ws[f"F{interest_row}"] = (
        f"=('Schedules'!F{interest_row - 1}+'Schedules'!F{closing_row})"
        f"/2*'Assumptions'!$C$45"
    )
    circular = tmp_path / "circular.xlsx"
    wb.save(circular)

    cycle = _find_cycle(_reference_graph(openpyxl.load_workbook(circular)))
    assert cycle is not None, "the finder missed a deliberately circular workbook"
    assert ("Schedules", f"F{interest_row}") in cycle
