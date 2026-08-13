"""Structure, layout and live-recalculation tests for the generated workbook.

Layout assertions here use LITERAL cell addresses. A test that resolves a row
through `layout.row_of` passes whatever position the writer chose, so it
cannot catch a positional error — which is exactly how a two-row title offset
survived Task 11's own test. Where this file asserts a position, it hardcodes
it.
"""

import re
from dataclasses import dataclass, replace
from pathlib import Path
from statistics import median

import openpyxl
import pytest

from bluebook.assumptions import Drivers, SCENARIOS
from bluebook.comps import (
    FY2025_NET_IMPAIRMENT_PPE,
    FY2025_NET_IMPAIRMENT_ROU,
    GREGGS_52_WEEK_HIGH,
    GREGGS_52_WEEK_LOW,
    GREGGS_SHARE_PRICE,
    GREGGS_SHARES_OUTSTANDING,
    PEERS,
    _TIE_TOLERANCE as TIE_TOLERANCE,
    greggs_trading_multiples,
    implied_value_from_comps,
    multiples,
)
from bluebook.inputs.greggs import GREGGS_HISTORICALS, GREGGS_SHARE_COUNT
from bluebook.lbo import SPONSOR_ENTRY_LEVERAGE, SPONSOR_IRR_HURDLE, greggs_lbo_case
from bluebook.recalc import recalc_values
from bluebook.reference import Model, build_model
from bluebook.valuation import (
    Valuation,
    enterprise_value,
    equity_bridge,
    excess_asset_tax_shield,
    implied_share_price,
    terminal_value_gordon,
    terminal_year,
    unlevered_fcf,
    value_model,
)
from bluebook.workbook.build import build_workbook
from bluebook.workbook.styles import HARDCODE_ALLOWED

FORECAST_COLS = "FGHIJ"

# Written out rather than imported from the sheet writers. A test that asked
# `sheet_comps` which column SSP Group landed in, or `sheet_sensitivity` where
# its grid starts, would pass whatever position those modules chose — which is
# the whole reason this file hardcodes addresses.
PEER_COLUMNS = ("C", "D", "E", "F", "G")
STAT_COLUMNS = {"min": "I", "median": "J", "max": "K"}
GRID_COLUMNS = ("D", "E", "F", "G", "H")
PRICE_GRID_ROWS = (5, 6, 7, 8, 9)
MULTIPLE_GRID_ROWS = (12, 13, 14, 15, 16)
GROWTH_AXIS_ROW = 4
# The axis offsets `sheet_sensitivity` applies, in the same order.
WACC_OFFSETS = (-0.01, -0.005, 0.0, 0.005, 0.01)
GROWTH_OFFSETS = (-0.005, -0.0025, 0.0, 0.0025, 0.005)


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory) -> Path:
    """One Base-case workbook, shared — building it is not what is under test."""
    path = tmp_path_factory.mktemp("wb") / "model.xlsx"
    return build_workbook(GREGGS_HISTORICALS, "Base", path)


@pytest.fixture(scope="module")
def workbook(workbook_path):
    return openpyxl.load_workbook(workbook_path)


# --------------------------------------------------------------------------
# Layout: literal addresses, per the plan's global constraints
# --------------------------------------------------------------------------


def test_every_sheet_puts_its_title_in_a1(workbook):
    for name in workbook.sheetnames:
        assert workbook[name]["A1"].value, f"{name} has no title in A1"


def test_year_headers_are_on_row_2_in_the_year_columns(workbook):
    assert workbook["Historicals"]["C2"].value == "FY2023"
    assert workbook["Historicals"]["D2"].value == "FY2024"
    assert workbook["Historicals"]["E2"].value == "FY2025"
    assert workbook["Assumptions"]["F2"].value == "FY2026"
    assert workbook["Assumptions"]["J2"].value == "FY2030"
    # The statements do not hold their own copy of the year labels.
    assert workbook["IS"]["F2"].value == "='Assumptions'!F$2"
    assert workbook["BS"]["J2"].value == "='Assumptions'!J$2"


def test_first_data_row_is_row_3_on_every_written_sheet(workbook):
    expected_first_labels = {
        "Assumptions": "SCENARIO (Bear / Base / Bull)",
        "Historicals": "Revenue",
        "IS": "Revenue",
        "BS": "Property, plant and equipment",
        "CF": "EBITDA",
        "Schedules": "Inventories (cost of sales x inventory days / 365)",
    }
    for sheet, label in expected_first_labels.items():
        assert workbook[sheet]["B3"].value == label, f"{sheet}!B3"


def test_the_cover_states_every_disclosure_it_must(workbook):
    """The five things a reader must learn without opening another tab.

    Three of them were missing and nobody noticed, because the only test of
    the Cover's content was the plan's own — which asserted that the strings
    "IFRS 16" and "net debt" appeared somewhere on the sheet, and passed on a
    Cover that said neither of the things those phrases were standing in for.
    The README meanwhile stated twice that the peer-provenance asymmetry was
    disclosed here, which it was not.

    Each assertion below names a substantive phrase from the disclosure it
    guards, not a keyword that could appear by accident in another note.
    """
    notes = " ".join(
        cell.value for cell in workbook["Cover"]["B"]
        if isinstance(cell.value, str)
    )

    required = {
        "lease treatment": "right-of-use depreciation sits inside D&A",
        "leases in the bridge": "deducted as debt in the equity bridge",
        "circularity decision": "interest on borrowings is charged on the opening balance",
        "RCF upsize, on the lease-inclusive basis": "Lease-inclusive net debt peaks",
        "the flattering basis, named as such": "gross borrowings alone would read",
        "peer provenance": "PROVENANCE IS NOT SYMMETRIC",
        "peer figures carry no page references": "carry no page references",
        "impairment convention": "IMPAIRMENT CONVENTION",
        "impairment is not like for like": "not like for like",
        "unsourced inputs named": "reasoned judgement estimates",
        "market data is an observation": "one-day",
        "drivers are not guidance": "not company guidance",
    }
    missing = [name for name, phrase in required.items() if phrase not in notes]
    assert not missing, (
        "the Cover does not disclose: " + ", ".join(missing)
    )


def test_the_cover_wraps_its_notes_and_is_the_one_unfrozen_sheet(workbook):
    """The Cover reads as prose; every other sheet keeps its frozen split.

    A frozen split at column C clips whatever column B cannot display, and the
    Cover's notes run to several hundred characters. Freezing bought that sheet
    nothing anyway — it has no year header and no grid to hold in view — so it
    is the one sheet written unfrozen and wrapped. The explicit row heights are
    part of the fix, not decoration: openpyxl writes no cached row heights, so
    a wrapped cell left to auto-fit opens one line high.
    """
    cover = workbook["Cover"]
    assert cover.freeze_panes is None

    notes = [cell for cell in cover["B"] if isinstance(cell.value, str) and cell.value]
    assert len(notes) >= 5, f"expected the Cover's disclosures, found {len(notes)}"
    for cell in notes:
        assert cell.alignment.wrap_text, f"Cover!B{cell.row} does not wrap"
        height = cover.row_dimensions[cell.row].height
        assert height and height >= 15.0, (
            f"Cover!B{cell.row} holds {len(cell.value)} characters but its row "
            f"height is {height!r}"
        )

    for name in workbook.sheetnames:
        if name != "Cover":
            assert workbook[name].freeze_panes == "C3", f"{name} is not frozen at C3"


def test_the_scenario_switch_sits_in_c3_and_holds_the_requested_scenario(workbook):
    assert workbook["Assumptions"]["C3"].value == "Base"


def test_the_scenario_switch_is_a_validated_dropdown_over_the_three_scenarios(workbook):
    validations = workbook["Assumptions"].data_validations.dataValidation
    covering = [dv for dv in validations if "C3" in str(dv.sqref)]
    assert len(covering) == 1, f"expected exactly one validation on C3, got {validations}"
    assert covering[0].type == "list"
    assert covering[0].formula1 == '"Bear,Base,Bull"'


def test_historicals_carries_the_reported_figures_and_their_sources(workbook):
    ws = workbook["Historicals"]
    assert ws["C3"].value == pytest.approx(GREGGS_HISTORICALS[0].revenue.value)
    assert ws["E3"].value == pytest.approx(GREGGS_HISTORICALS[-1].revenue.value)
    assert "FY2023 AR p.119" in ws["L3"].value


# --------------------------------------------------------------------------
# The rule the whole task exists for: formulas, not values
# --------------------------------------------------------------------------


@pytest.mark.parametrize("sheet", ["IS", "BS", "CF", "Schedules", "DCF", "Sensitivity", "LBO"])
def test_no_forecast_cell_on_a_calculation_sheet_is_a_constant(workbook, sheet):
    ws = workbook[sheet]
    for row in ws.iter_rows(min_col=6, max_col=10):
        for cell in row:
            if cell.value is not None:
                assert isinstance(cell.value, str) and cell.value.startswith("="), (
                    f"{sheet}!{cell.coordinate} holds a constant: {cell.value!r}"
                )


def test_calculation_sheets_hold_no_hardcoded_number_anywhere(workbook):
    """Not just F:J — no numeric constant in any cell of a calculation sheet."""
    for name in workbook.sheetnames:
        if name in HARDCODE_ALLOWED:
            continue
        for row in workbook[name].iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), (
                    f"{name}!{cell.coordinate} is a hardcoded number: {cell.value!r}"
                )


def test_driver_rows_choose_between_the_three_scenario_paths(workbook):
    ws = workbook["Assumptions"]
    revenue_growth = next(
        r for r in range(3, ws.max_row + 1) if ws[f"B{r}"].value == "Revenue growth"
    )
    formula = ws[f"F{revenue_growth}"].value
    assert formula.startswith('=CHOOSE(MATCH(\'Assumptions\'!$C$3,{"Bear";"Base";"Bull"},0),')
    # Three arms, one per scenario, all reading column F.
    assert formula.count("'Assumptions'!F") == 3


# --------------------------------------------------------------------------
# Iterative calculation
# --------------------------------------------------------------------------


def test_iterative_calculation_is_enabled_with_the_spikes_settings(workbook):
    """Still pinned even though the workbook is now acyclic.

    The settings are retained on purpose — free on a file with no cycle, and
    they turn a future reintroduced circularity into a converged answer rather
    than `Err:522`. See `build.py`'s docstring. Dropping them would be a silent
    change to how such an edit behaves, so it should have to fail a test.
    """
    assert workbook.calculation.iterate is True
    assert workbook.calculation.iterateCount == 100
    assert workbook.calculation.iterateDelta == pytest.approx(0.0001)


# --------------------------------------------------------------------------
# The build's own guard
# --------------------------------------------------------------------------


def test_build_rejects_an_unknown_scenario(tmp_path):
    with pytest.raises(ValueError, match="scenario_name must be one of"):
        build_workbook(GREGGS_HISTORICALS, "Sideways", tmp_path / "x.xlsx")


def test_build_rejects_an_empty_history(tmp_path):
    with pytest.raises(ValueError, match="at least one reported year"):
        build_workbook([], "Base", tmp_path / "x.xlsx")


# --------------------------------------------------------------------------
# Live recalculation: does Excel agree with the Python model, and does the
# switch re-drive it?
# --------------------------------------------------------------------------

# Every row of the model that `reference.py` also computes, checked against it
# after a real LibreOffice recalculation. The financing block is included: since
# `INTEREST_BASIS` became "opening" the workbook has no circular reference, so
# there is nothing here that LibreOffice cannot resolve and no reason to hold
# any row back. Both balance sheet totals and the balance check are included
# too, which is what makes this a test of the model rather than of a subset.
CHECKED_ROWS = (
    ("IS", "Revenue", lambda m: m.income_statement["revenue"]),
    ("IS", "Gross profit", lambda m: m.income_statement["gross_profit"]),
    ("IS", "Cost of sales", lambda m: m.income_statement["cost_of_sales"]),
    ("IS", "Operating costs", lambda m: m.income_statement["operating_costs"]),
    ("IS", "EBITDA", lambda m: m.ebitda),
    ("IS", "Depreciation of PP&E", lambda m: m.income_statement["depreciation_ppe"]),
    ("IS", "Depreciation of ROU assets", lambda m: m.income_statement["depreciation_rou"]),
    ("IS", "Amortisation", lambda m: m.income_statement["amortisation"]),
    ("IS", "Total depreciation and amortisation", lambda m: m.da_total),
    ("IS", "EBIT", lambda m: m.ebit),
    ("IS", "Lease interest", lambda m: m.income_statement["lease_interest"]),
    ("IS", "Interest on borrowings", lambda m: m.income_statement["debt_interest"]),
    ("IS", "Profit before tax", lambda m: m.income_statement["profit_before_tax"]),
    ("IS", "Tax charge", lambda m: m.income_statement["tax"]),
    ("IS", "Net income", lambda m: m.net_income),
    ("BS", "Property, plant and equipment", lambda m: m.balance_sheet["ppe"]),
    ("BS", "Right-of-use assets", lambda m: m.balance_sheet["rou_assets"]),
    ("BS", "Intangible assets", lambda m: m.balance_sheet["intangibles"]),
    ("BS", "Inventories", lambda m: m.balance_sheet["inventories"]),
    ("BS", "Trade and other receivables", lambda m: m.balance_sheet["trade_receivables"]),
    ("BS", "Cash and cash equivalents", lambda m: m.balance_sheet["cash"]),
    ("BS", "Other assets", lambda m: m.balance_sheet["other_assets"]),
    ("BS", "Total assets", lambda m: m.balance_sheet["total_assets"]),
    ("BS", "Trade and other payables", lambda m: m.balance_sheet["trade_payables"]),
    ("BS", "Lease liabilities", lambda m: m.balance_sheet["lease_liabilities"]),
    ("BS", "Borrowings", lambda m: m.balance_sheet["borrowings"]),
    ("BS", "Other liabilities", lambda m: m.balance_sheet["other_liabilities"]),
    ("BS", "Total equity", lambda m: m.balance_sheet["equity"]),
    (
        "BS",
        "Total liabilities and equity",
        lambda m: m.balance_sheet["total_liabilities_and_equity"],
    ),
    ("BS", "Balance check", lambda m: [0.0] * len(m.years)),
    ("CF", "EBITDA", lambda m: m.cash_flow["ebitda"]),
    (
        "CF",
        "Increase in working capital",
        lambda m: m.cash_flow["change_in_working_capital"],
    ),
    ("CF", "Tax paid", lambda m: m.cash_flow["tax_paid"]),
    ("CF", "Cash from operations", lambda m: m.cash_flow["cash_from_operations"]),
    ("CF", "Capital expenditure", lambda m: m.cash_flow["capex"]),
    ("CF", "of which PP&E", lambda m: m.cash_flow["capex_ppe"]),
    ("CF", "of which intangible", lambda m: m.cash_flow["capex_intangible"]),
    ("CF", "Lease interest paid", lambda m: m.cash_flow["lease_interest_paid"]),
    ("CF", "Lease principal repaid", lambda m: m.cash_flow["lease_principal_paid"]),
    ("CF", "Interest paid on borrowings", lambda m: m.cash_flow["debt_interest_paid"]),
    ("CF", "Repayment of borrowings", lambda m: m.cash_flow["debt_repayment"]),
    ("CF", "Revolver draw", lambda m: m.cash_flow["revolver_draw"]),
    ("CF", "Dividends paid", lambda m: m.cash_flow["dividends_paid"]),
    ("CF", "Net change in cash", lambda m: m.cash_flow["net_change_in_cash"]),
    ("CF", "Opening cash", lambda m: m.cash_flow["opening_cash"]),
    ("CF", "Closing cash", lambda m: m.cash_flow["closing_cash"]),
    ("Schedules", "Inventories", lambda m: m.working_capital.inventories),
    ("Schedules", "Trade receivables", lambda m: m.working_capital.receivables),
    ("Schedules", "Trade payables", lambda m: m.working_capital.payables),
    (
        "Schedules",
        "Net working capital",
        lambda m: m.working_capital.net_working_capital,
    ),
    ("Schedules", "Increase in net working capital", lambda m: m.working_capital.change_in_nwc),
    ("Schedules", "PP&E capex", lambda m: m.fixed_assets.capex),
    ("Schedules", "Depreciation of PP&E", lambda m: m.fixed_assets.depreciation),
    ("Schedules", "Closing PP&E", lambda m: m.fixed_assets.closing_ppe),
    ("Schedules", "Intangible additions", lambda m: m.cash_flow["capex_intangible"]),
    ("Schedules", "Amortisation", lambda m: m.income_statement["amortisation"]),
    ("Schedules", "Closing intangibles", lambda m: m.balance_sheet["intangibles"]),
    ("Schedules", "Total cash capex", lambda m: m.cash_flow["capex"]),
    ("Schedules", "New ROU asset additions", lambda m: m.leases.additions),
    ("Schedules", "Depreciation of ROU assets", lambda m: m.leases.depreciation),
    ("Schedules", "Closing right-of-use assets", lambda m: m.leases.closing_rou),
    ("Schedules", "Lease interest", lambda m: m.leases.interest),
    ("Schedules", "Lease principal repaid", lambda m: m.leases.principal_paid),
    ("Schedules", "Closing lease liabilities", lambda m: m.leases.closing_liability),
    ("Schedules", "Dividends declared", lambda m: m.cash_flow["dividends_paid"]),
    ("Schedules", "Closing shareholders' equity", lambda m: m.balance_sheet["equity"]),
    ("Schedules", "Opening borrowings", lambda m: m.debt.opening),
    ("Schedules", "Interest on borrowings", lambda m: m.debt.interest),
    ("Schedules", "Repayment of borrowings", lambda m: m.debt.repayment),
    ("Schedules", "Revolver draw", lambda m: m.debt.revolver_draw),
    ("Schedules", "Closing borrowings", lambda m: m.debt.closing),
    ("Schedules", "Closing cash", lambda m: m.debt.cash_balance),
)


# --------------------------------------------------------------------------
# The valuation sheets, checked the same way against valuation.py, comps.py
# and lbo.py — the three modules that own the numbers on them
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class _Expected:
    """Everything the four valuation sheets are checked against, for one scenario.

    Assembled from the model modules and nothing else: the point of the
    comparison is that the workbook's formulas reproduce `valuation.py`,
    `comps.py` and `lbo.py`, so anything computed here has to come from them.
    """

    drivers: Drivers
    model: Model
    valuation: Valuation
    lbo: object
    peer: dict[str, dict[str, float]]
    greggs: object


def _expected_for(scenario: str) -> _Expected:
    drivers = SCENARIOS[scenario]
    model = build_model(GREGGS_HISTORICALS, drivers)
    greggs = greggs_trading_multiples(GREGGS_HISTORICALS)
    return _Expected(
        drivers=drivers,
        model=model,
        # The DCF divides by the FILING's weighted average diluted count, not
        # the price provider's shares-outstanding figure — they differ by 0.51%
        # and the workbook carries both.
        valuation=value_model(
            model, drivers, GREGGS_HISTORICALS, GREGGS_SHARE_COUNT.value
        ),
        # The sponsor pays the market: entry EV is Greggs' own traded EV.
        lbo=greggs_lbo_case(model, drivers, GREGGS_HISTORICALS, greggs.ev),
        peer=multiples(PEERS),
        greggs=greggs,
    )


def _lbo_opening_debt(e: _Expected) -> list[float]:
    """Opening financial debt: the entry draw, then each prior year's close."""
    return [e.lbo.entry_financial_debt] + e.lbo.financial_debt_closing[:-1]


def _lbo_total_debt(e: _Expected) -> list[float]:
    return [
        financial + leases
        for financial, leases in zip(
            e.lbo.financial_debt_closing, e.model.leases.closing_liability
        )
    ]


# Per-year rows on the DCF and LBO sheets. Same shape as CHECKED_ROWS above,
# but the expectation comes from `valuation.py` / `lbo.py` rather than from
# `reference.py`, so the lambdas take the bundle rather than the model.
VALUATION_ROWS = (
    ("DCF", "EBIT (from the income statement)", lambda e: e.model.ebit),
    (
        "DCF",
        "EBIT after tax",
        lambda e: [ebit * (1.0 - e.drivers.tax_rate) for ebit in e.model.ebit],
    ),
    ("DCF", "Depreciation and amortisation", lambda e: e.model.da_total),
    ("DCF", "Total cash capex", lambda e: e.model.cash_flow["capex"]),
    ("DCF", "New ROU asset additions", lambda e: e.model.leases.additions),
    ("DCF", "Increase in net working capital", lambda e: e.model.working_capital.change_in_nwc),
    ("DCF", "Unlevered free cash flow", lambda e: e.valuation.fcf),
    # The mid-year convention, which the sheet builds from COLUMN() rather than
    # writing out five numbers. If that expression ever slipped a column, every
    # discount factor on the sheet would be off and this row is what says so.
    ("DCF", "Discount period", lambda e: [0.5, 1.5, 2.5, 3.5, 4.5]),
    ("DCF", "Discount factor", lambda e: e.valuation.discount_factors),
    ("DCF", "PV of unlevered free cash flow", lambda e: e.valuation.pv_fcf),
    ("LBO", "Unlevered free cash flow", lambda e: e.valuation.fcf),
    ("LBO", "New ROU additions added back", lambda e: e.model.leases.additions),
    ("LBO", "Lease principal repaid", lambda e: e.model.leases.principal_paid),
    ("LBO", "Lease interest", lambda e: e.model.leases.interest),
    ("LBO", "Opening financial debt", _lbo_opening_debt),
    (
        "LBO",
        "Interest on financial debt",
        lambda e: [e.lbo.interest_rate * debt for debt in _lbo_opening_debt(e)],
    ),
    ("LBO", "Cash available to repay debt", lambda e: e.lbo.cash_swept),
    ("LBO", "Closing financial debt", lambda e: e.lbo.financial_debt_closing),
    ("LBO", "Closing lease liabilities", lambda e: e.model.leases.closing_liability),
    ("LBO", "Closing total debt including leases", _lbo_total_debt),
    (
        "LBO",
        "Closing total debt / that year's EBITDA",
        lambda e: [
            debt / ebitda for debt, ebitda in zip(_lbo_total_debt(e), e.model.ebitda)
        ],
    ),
)

# Single cells. Every scalar output of the three modules that reaches a sheet,
# so the headline figures an interviewer reads off the workbook — the WACC, the
# implied share price, the money multiple, the IRR — are inside the comparison
# rather than beside it.
VALUATION_CELLS = (
    ("DCF", "Cost of equity", "C", lambda e: e.valuation.cost_of_equity),
    ("DCF", "After-tax cost of debt", "C", lambda e: e.valuation.after_tax_cost_of_debt),
    ("DCF", "WACC", "C", lambda e: e.valuation.wacc),
    ("DCF", "Perpetuity growth rate (g*)", "C", lambda e: e.drivers.perpetuity_growth),
    ("DCF", "Terminal PP&E / revenue", "C", lambda e: e.valuation.terminal.ppe_intensity),
    ("DCF", "Terminal ROU assets / revenue", "C", lambda e: e.valuation.terminal.rou_intensity),
    (
        "DCF",
        "Terminal total capex",
        "C",
        lambda e: e.valuation.terminal.capex_pct_revenue,
    ),
    (
        "DCF",
        "Terminal ROU additions",
        "C",
        lambda e: e.valuation.terminal.rou_additions_pct_revenue,
    ),
    ("DCF", "Terminal D&A (% of revenue)", "C", lambda e: e.valuation.terminal.da_pct_revenue),
    (
        "DCF",
        "Terminal net working capital",
        "C",
        lambda e: e.valuation.terminal.nwc_pct_revenue,
    ),
    ("DCF", "Terminal revenue", "C", lambda e: e.valuation.terminal.revenue),
    ("DCF", "Terminal EBITDA margin", "C", lambda e: e.valuation.terminal.ebitda_margin),
    ("DCF", "Terminal EBITDA (margin x revenue)", "C", lambda e: e.valuation.terminal.ebitda),
    ("DCF", "Terminal depreciation and amortisation", "C", lambda e: e.valuation.terminal.da),
    ("DCF", "Terminal EBIT", "C", lambda e: e.valuation.terminal.ebit),
    ("DCF", "Terminal total cash capex", "C", lambda e: e.valuation.terminal.capex),
    ("DCF", "Terminal ROU asset additions", "C", lambda e: e.valuation.terminal.rou_additions),
    (
        "DCF",
        "Terminal increase in net working capital",
        "C",
        lambda e: e.valuation.terminal.change_in_nwc,
    ),
    ("DCF", "Terminal unlevered free cash flow", "C", lambda e: e.valuation.terminal.fcf),
    (
        "DCF",
        "Terminal D&A / EBITDA",
        "C",
        lambda e: e.valuation.terminal.da / e.valuation.terminal.ebitda,
    ),
    ("DCF", "Terminal value (Gordon growth)", "C", lambda e: e.valuation.terminal_value_gordon),
    ("DCF", "Excess PP&E over the re-based base, at FY2029", "C", lambda e: e.valuation.excess_ppe),
    (
        "DCF",
        "Value at the terminal date of the excess PP&E tax shield",
        "C",
        lambda e: e.valuation.excess_ppe_tax_shield,
    ),
    (
        "DCF",
        "Terminal value (exit multiple)",
        "C",
        lambda e: e.valuation.terminal_value_exit_multiple,
    ),
    # The identity the caveat on the sheet rests on: the exit-multiple terminal
    # value IS the peer median EV/EBIT times terminal EBIT, to within the
    # two-decimal rounding of the shipped driver.
    (
        "DCF",
        "Memo: exit-multiple TV / terminal EBIT",
        "C",
        lambda e: e.valuation.terminal_value_exit_multiple / e.valuation.terminal.ebit,
    ),
    ("DCF", "Memo: peer median EV/EBIT from Comps", "C", lambda e: e.peer["ev_ebit"]["median"]),
    (
        "DCF",
        "Gordon terminal value / terminal EBITDA",
        "C",
        lambda e: e.valuation.terminal_value_gordon / e.valuation.terminal.ebitda,
    ),
    (
        "DCF",
        "Gordon terminal value / terminal EBIT",
        "C",
        lambda e: e.valuation.terminal_value_gordon / e.valuation.terminal.ebit,
    ),
    ("DCF", "PV of the explicit forecast period", "C", lambda e: sum(e.valuation.pv_fcf)),
    ("DCF", "PV of the terminal value", "C", lambda e: e.valuation.pv_terminal_value),
    ("DCF", "Enterprise value", "C", lambda e: e.valuation.enterprise_value),
    (
        "DCF",
        "Terminal value as a share of enterprise value",
        "C",
        lambda e: e.valuation.pv_terminal_value / e.valuation.enterprise_value,
    ),
    ("DCF", "Net debt at FY2025", "C", lambda e: e.valuation.net_debt),
    ("DCF", "Lease liabilities at FY2025", "C", lambda e: e.valuation.lease_liabilities),
    ("DCF", "Equity value", "C", lambda e: e.valuation.equity_value),
    ("DCF", "Shares in issue", "C", lambda e: GREGGS_SHARE_COUNT.value),
    ("DCF", "Implied share price (p)", "C", lambda e: e.valuation.share_price_pence),
    # --- Comps: Greggs on the peers' basis ---
    ("Comps", "Greggs market capitalisation", "C", lambda e: e.greggs.market_cap),
    ("Comps", "Greggs net debt including lease liabilities", "C",
     lambda e: e.greggs.net_debt_incl_leases),
    ("Comps", "Greggs enterprise value", "C", lambda e: e.greggs.ev),
    ("Comps", "Greggs EBITDA", "C", lambda e: e.greggs.ebitda),
    (
        "Comps",
        "Greggs EBIT (£m, FY2025, struck AFTER impairment — see below)",
        "C",
        lambda e: e.greggs.ebit,
    ),
    ("Comps", "Greggs net income", "C", lambda e: e.greggs.net_income),
    ("Comps", "Greggs EV / EBITDA", "C", lambda e: e.greggs.ev_ebitda),
    ("Comps", "Greggs EV / EBIT", "C", lambda e: e.greggs.ev_ebit),
    ("Comps", "Greggs P/E", "C", lambda e: e.greggs.pe),
    ("Comps", "FY2025 net impairment inside D&A", "C", lambda e: e.greggs.impairment),
    (
        "Comps",
        "Greggs EBIT on the peers' basis, before impairment (£m)",
        "C",
        lambda e: e.greggs.ebit_pre_impairment,
    ),
    (
        "Comps",
        "Peers'-basis EV / EBIT (Greggs, before impairment)",
        "C",
        lambda e: e.greggs.ev_ebit_pre_impairment,
    ),
    # The exit multiple this peer set implies, and the rounding the shipped
    # driver carries against it.
    (
        "Comps",
        "Terminal D&A / EBITDA, from the DCF's re-based terminal year",
        "C",
        lambda e: e.valuation.terminal.da / e.valuation.terminal.ebitda,
    ),
    (
        "Comps",
        "Derived exit EV/EBITDA",
        "C",
        lambda e: e.peer["ev_ebit"]["median"]
        * (1.0 - e.valuation.terminal.da / e.valuation.terminal.ebitda),
    ),
    ("Comps", "Exit EV/EBITDA as shipped on Assumptions", "C",
     lambda e: e.drivers.exit_ev_ebitda),
    # --- LBO ---
    ("LBO", "Entry enterprise value", "C", lambda e: e.lbo.entry_ev),
    ("LBO", "Entry EBITDA", "C", lambda e: e.lbo.entry_ebitda),
    ("LBO", "Entry EV/EBITDA", "C", lambda e: e.lbo.entry_multiple),
    ("LBO", "Entry leverage", "C", lambda e: float(SPONSOR_ENTRY_LEVERAGE)),
    ("LBO", "Entry total net debt including leases", "C", lambda e: e.lbo.entry_debt),
    ("LBO", "new financial debt raised", "C", lambda e: e.lbo.entry_financial_debt),
    ("LBO", "Sponsor equity cheque", "C", lambda e: e.lbo.returns.entry_equity),
    ("LBO", "Exit EBITDA", "C", lambda e: e.lbo.exit_ebitda),
    (
        "LBO",
        "Exit EV/EBITDA (the entry multiple, unchanged)",
        "C",
        lambda e: e.lbo.exit_multiple,
    ),
    ("LBO", "Exit enterprise value", "C", lambda e: e.lbo.exit_ev),
    ("LBO", "Exit total net debt including leases", "C", lambda e: e.lbo.exit_debt),
    ("LBO", "Exit equity value", "C", lambda e: e.lbo.returns.exit_equity),
    ("LBO", "Hold period", "C", lambda e: float(e.lbo.years)),
    ("LBO", "Money multiple", "C", lambda e: e.lbo.returns.money_multiple),
    ("LBO", "IRR (geometric: one cheque in, one out)", "C", lambda e: e.lbo.returns.irr),
    ("LBO", "IRR less the sponsor hurdle", "C", lambda e: e.lbo.returns.irr - SPONSOR_IRR_HURDLE),
    # --- Sensitivity: the centre of each grid is the live case ---
    (
        "Sensitivity",
        "WACC (the model's own)",
        "F",
        lambda e: e.valuation.share_price_pence,
    ),
    (
        "Sensitivity",
        "Implied terminal multiple at WACC (the model's own)",
        "F",
        lambda e: e.valuation.terminal_value_gordon / e.valuation.terminal.ebitda,
    ),
    (
        "Sensitivity",
        "Terminal unlevered free cash flow",
        "F",
        lambda e: e.valuation.terminal.fcf,
    ),
    ("Sensitivity", "Excess PP&E over this growth rate's", "F", lambda e: e.valuation.excess_ppe),
    ("Sensitivity", "Check: centre column's terminal FCF", "C", lambda e: 0.0),
)


def _peer_cells() -> tuple[tuple[str, str, str, float], ...]:
    """One expectation per peer per derived multiple, plus the statistics.

    Built from `comps.PEERS` and `comps.multiples` so the sheet is checked
    against the module that owns the peer set, but the COLUMNS are the literal
    letters above: a wrong column here is a wrong company, and resolving the
    column dynamically would agree with whatever the writer did.
    """
    cells: list[tuple[str, str, str, float]] = []
    derived: tuple[tuple[str, object], ...] = (
        ("EV / EBITDA", lambda p: p.ev / p.ebitda),
        ("EV / EBIT", lambda p: p.ev / p.ebit),
        ("P/E", lambda p: p.market_cap / p.net_income),
        ("D&A / EBITDA", lambda p: (p.ebitda - p.ebit) / p.ebitda),
        ("Implied D&A", lambda p: p.ebitda - p.ebit),
    )
    for label, value_of in derived:
        for peer, col in zip(PEERS, PEER_COLUMNS):
            cells.append(("Comps", label, col, value_of(peer)))
    for key, label in (("ev_ebitda", "EV / EBITDA"), ("ev_ebit", "EV / EBIT"), ("pe", "P/E")):
        for which, col in STAT_COLUMNS.items():
            cells.append(("Comps", label, col, multiples(PEERS)[key][which]))
    return tuple(cells)


PEER_CELLS = _peer_cells()


# The ten transcribed figures behind every peer multiple, checked cell by cell
# against the `Peer` records themselves.
#
# `PEER_CELLS` above does not reach these. It checks ratios, and a ratio pins
# only its own two operands: mis-transcribe a peer's share price, its shares
# outstanding, its lease liabilities or its minority interests and every
# derived multiple still agrees, because none of those four is a numerator or a
# denominator of any of them. The sheet's own tie rows would notice — which is
# why they are asserted too, three tuples down — but nothing asserted THEM
# either. Ten rows against the module's fields closes the whole family at once,
# and the expectation is the field, not a re-derivation of the sheet's formula.
PEER_INPUT_ROWS = (
    ("Share price (pence)", lambda p: p.share_price_pence),
    ("Shares outstanding (m)", lambda p: p.shares),
    ("Market capitalisation (£m)", lambda p: p.market_cap),
    ("Net debt including lease liabilities (£m)", lambda p: p.net_debt_incl_leases),
    ("of which lease liabilities (£m)", lambda p: p.lease_liabilities),
    ("Minority interests (£m, book value)", lambda p: p.minority_interests),
    ("Enterprise value (£m)", lambda p: p.ev),
    ("EBITDA (£m, post-IFRS 16, underlying)", lambda p: p.ebitda),
    ("EBIT (£m, post-IFRS 16, underlying)", lambda p: p.ebit),
    ("Net income (£m, statutory, attributable)", lambda p: p.net_income),
)


def _peer_input_cells() -> tuple[tuple[str, str, str, float], ...]:
    return tuple(
        ("Comps", label, col, value_of(peer))
        for label, value_of in PEER_INPUT_ROWS
        for peer, col in zip(PEERS, PEER_COLUMNS)
    )


PEER_INPUT_CELLS = _peer_input_cells()

# The two identities the peer table states about itself, one per peer. Both
# read zero by construction when the transcription is right, so they are worth
# asserting only because the rows above can now no longer drift silently: these
# catch a wrong EV or market cap on the SHEET even if the module agreed.
PEER_TIE_ROWS = (
    "Check: price x shares less market capitalisation (£m, |x| < 0.01)",
    "Check: market cap + net debt + minorities less EV (£m, |x| < 0.01)",
)

# Greggs' own market observations, and the peer capital-intensity statistics.
# The 52-week range is the one figure in the workbook that goes stale, so it is
# pinned to the sourced constant rather than left to be read off the sheet.
GREGGS_MARKET_CELLS = (
    # The PP&E/ROU split of FY2025 net impairment. Their SUM is checked
    # against `greggs.impairment`, which is why these were first left out as
    # "a transcription whose only honest expectation would be the sheet's own
    # literals" — wrong on both counts. `comps.py` names them as constants and
    # `sheet_comps.py` already imports them, so the honest expectation is the
    # module's, exactly as for the peer inputs. And the sum is not cover:
    # swapping 3.9 and 3.0 leaves it unchanged and escaped every test.
    ("Comps", "FY2025 net impairment of PP&E, inside depreciation", "C",
     lambda: FY2025_NET_IMPAIRMENT_PPE),
    ("Comps", "FY2025 net impairment of ROU assets, inside depreciation", "C",
     lambda: FY2025_NET_IMPAIRMENT_ROU),
    ("Comps", "Greggs share price (pence)", "C", lambda: GREGGS_SHARE_PRICE.value),
    ("Comps", "Greggs shares outstanding", "C", lambda: GREGGS_SHARES_OUTSTANDING.value),
    ("Comps", "52-week low", "C", lambda: GREGGS_52_WEEK_LOW.value),
    ("Comps", "52-week high", "C", lambda: GREGGS_52_WEEK_HIGH.value),
)


def _da_intensity_stat_cells() -> tuple[tuple[str, str, str, float], ...]:
    """Min/median/max of D&A/EBITDA across the peers, from `comps.PEERS`.

    `multiples()` does not carry this statistic, so it is computed here from
    the same peer records the sheet's MIN/MEDIAN/MAX range covers. That is the
    module's data rather than the sheet's arithmetic, which is the distinction
    that matters — the sheet computes it from its own five cells, and those
    five cells are pinned by `PEER_CELLS`.
    """
    intensities = [(p.ebitda - p.ebit) / p.ebitda for p in PEERS]
    stats = {
        "min": min(intensities),
        "median": median(intensities),
        "max": max(intensities),
    }
    return tuple(
        ("Comps", "D&A / EBITDA", STAT_COLUMNS[which], value)
        for which, value in stats.items()
    )


DA_INTENSITY_STAT_CELLS = _da_intensity_stat_cells()


def _comps_implied(e: _Expected, which: str) -> float:
    """The comps-implied share price at one end of the peer EV/EBITDA range.

    Composed from `comps.implied_value_from_comps`, the function that owns this
    arithmetic, rather than restated here. Net debt is passed EXCLUDING leases
    with the lease liability separately, which is that function's contract and
    the same split `valuation.equity_bridge` makes.
    """
    last = GREGGS_HISTORICALS[-1]
    return implied_value_from_comps(
        e.greggs.ebitda,
        e.peer["ev_ebitda"][which],
        last.borrowings.value - last.cash.value,
        last.lease_liabilities.value,
        GREGGS_SHARE_COUNT.value,
    )


# Leaf cells: outputs with nothing downstream of them.
#
# These are the reason a coverage audit was worth running. Every other
# unchecked cell on a calculation sheet is a working that feeds a checked one,
# so a mutation to it surfaces somewhere in the comparison. These do not feed
# anything — Comps C47:C49 are what the football field's comps bar draws, and
# the discounts and the LBO leaves are figures a reader quotes straight off the
# sheet — so before this table nothing in the suite would have failed if any of
# them were wrong.
#
# That was measured too, against the commit before this one: of twelve cells
# this task added checks for, eleven were mutated by 1% without a single test
# failing. Only the peers' market capitalisation was already pinned, and only
# because it is a denominator of P/E. The largest family was the peer table
# itself — mis-state a peer's share price, shares outstanding, lease
# liabilities or minority interests and every derived multiple still agreed.
LEAF_CELLS = (
    (
        "Comps",
        "Comps-implied share price at the peer minimum EV/EBITDA",
        "C",
        lambda e: _comps_implied(e, "min"),
    ),
    (
        "Comps",
        "Comps-implied share price at the peer median EV/EBITDA",
        "C",
        lambda e: _comps_implied(e, "median"),
    ),
    (
        "Comps",
        "Comps-implied share price at the peer maximum EV/EBITDA",
        "C",
        lambda e: _comps_implied(e, "max"),
    ),
    (
        "Comps",
        "Discount to the peer median on EV/EBITDA",
        "C",
        lambda e: e.greggs.ev_ebitda / e.peer["ev_ebitda"]["median"] - 1.0,
    ),
    (
        "Comps",
        "Discount to the peer median on EV/EBIT",
        "C",
        lambda e: e.greggs.ev_ebit / e.peer["ev_ebit"]["median"] - 1.0,
    ),
    (
        "Comps",
        "Discount to the peer median on the peers'-basis EV/EBIT",
        "C",
        lambda e: e.greggs.ev_ebit_pre_impairment / e.peer["ev_ebit"]["median"] - 1.0,
    ),
    (
        "Comps",
        "Rounding carried by the shipped driver",
        "C",
        lambda e: e.peer["ev_ebit"]["median"]
        * (1.0 - e.valuation.terminal.da / e.valuation.terminal.ebitda)
        - e.drivers.exit_ev_ebitda,
    ),
    (
        "DCF",
        "Implied share price against the traded price on Comps",
        "C",
        lambda e: e.valuation.share_price_pence / GREGGS_SHARE_PRICE.value - 1.0,
    ),
    # Sensitivity's two "for comparison" links. They restate a Comps cell for
    # side-by-side reading and feed nothing, so despite sitting among the
    # sensitivity workings they are leaves: point either at the wrong Comps
    # cell and the reader sees a wrong number with every total still right.
    (
        "Sensitivity",
        "For comparison: peer median EV/EBITDA",
        "C",
        lambda e: e.peer["ev_ebitda"]["median"],
    ),
    (
        "Sensitivity",
        "For comparison: exit EV/EBITDA derived from the peer median EV/EBIT",
        "C",
        lambda e: e.peer["ev_ebit"]["median"]
        * (1.0 - e.valuation.terminal.da / e.valuation.terminal.ebitda),
    ),
    ("LBO", "of which the existing lease liability", "C", lambda e: e.lbo.entry_leases),
    (
        "LBO",
        "Change in total debt over the hold",
        "C",
        lambda e: e.lbo.exit_debt - e.lbo.entry_debt,
    ),
    (
        "LBO",
        "EBITDA growth over the hold",
        "C",
        lambda e: e.lbo.exit_ebitda / e.lbo.entry_ebitda - 1.0,
    ),
    # The exit multiple at which the sponsor would just clear its hurdle.
    # `lbo.py` exposes no field for it, so this states the DEFINITION of the
    # quantity from that module's own outputs — the exit EV that turns the
    # entry cheque into a hurdle-rate return, over exit EBITDA — rather than
    # copying the sheet's formula, which is the same identity arranged
    # differently and would agree with the sheet however the sheet was wrong.
    (
        "LBO",
        "Exit EV/EBITDA that would clear the hurdle",
        "C",
        lambda e: (
            e.lbo.returns.entry_equity * (1.0 + SPONSOR_IRR_HURDLE) ** e.lbo.years
            + e.lbo.exit_debt
        )
        / e.lbo.exit_ebitda,
    ),
)


# The football field's three bars, in the units the chart plots. Nothing else
# checks these numerically: `test_football_field.py` pins the sheet's structure
# — three bars, no exit-multiple bar, the span on its own row — and the chart's
# wiring, but never asked what the bars are worth.
FOOTBALL_FIELD_BARS = (
    (
        "DCF (Gordon)",
        lambda e: min(
            _sensitivity_price(
                e, e.valuation.wacc + w, e.drivers.perpetuity_growth + g
            )
            for w in WACC_OFFSETS
            for g in GROWTH_OFFSETS
        ),
        lambda e: max(
            _sensitivity_price(
                e, e.valuation.wacc + w, e.drivers.perpetuity_growth + g
            )
            for w in WACC_OFFSETS
            for g in GROWTH_OFFSETS
        ),
        lambda e: e.valuation.share_price_pence,
    ),
    (
        "Trading comps",
        lambda e: _comps_implied(e, "min"),
        lambda e: _comps_implied(e, "max"),
        lambda e: _comps_implied(e, "median"),
    ),
    (
        "52-week traded range",
        lambda e: GREGGS_52_WEEK_LOW.value,
        lambda e: GREGGS_52_WEEK_HIGH.value,
        lambda e: GREGGS_SHARE_PRICE.value,
    ),
)


# The workbook is acyclic, so LibreOffice evaluates every cell in one pass and
# the only difference from Python is floating-point association. 1e-9 is
# therefore a real assertion rather than a nod: under the old average-interest
# basis the tolerance had to leave room for iterative convergence slack
# (~1e-4), and it no longer does.
RECALC_TOLERANCE = 1e-9

# Every error string either engine can leave in a cell: Excel's seven, and
# LibreOffice's own `Err:NNN` form (522 is its circular-reference code, which is
# the one this workbook's design is meant to make impossible).
ERROR_VALUE = re.compile(r"^(#(REF!|DIV/0!|VALUE!|NAME\?|N/A|NULL!|NUM!)|Err:\d+)$")


def _row_of(ws, label_prefix: str) -> int:
    """The single row on `ws` whose label starts with `label_prefix`.

    Raises if two rows match, because a prefix that matches more than one row
    silently checks the wrong one — "EBIT" matching "EBITDA" cost an hour of
    chasing a phantom defect during this task.
    """
    matches = [
        row
        for row in range(3, ws.max_row + 1)
        # Leading spaces indent the cash flow's memo lines under their total;
        # they are presentation, so they are not part of the label here.
        if (label := (ws[f"B{row}"].value or "").lstrip())
        and (label == label_prefix or label.startswith(label_prefix + " "))
    ]
    assert len(matches) == 1, (
        f"{ws.title}: {label_prefix!r} matches {len(matches)} rows {matches}; "
        f"a prefix must identify exactly one row"
    )
    return matches[0]


def _compare_against_reference(path: Path, scenario: str) -> tuple[float, set[str]]:
    """Recalculate `path` and check every checked row and cell.

    Returns the worst absolute difference and the set of `"Sheet!ADDR"` cells
    this comparison actually asserted. The second half is what
    `test_every_computed_cell_on_a_calculation_sheet_is_inside_the_cross_check`
    audits: coverage claimed by counting entries in the tables above would
    miss the fifty sensitivity-grid cells (checked at literal addresses, not by
    label) and would not notice a row added to a sheet and to no table.
    Recording the addresses as they are checked is the only account of
    coverage that cannot drift from the checking.

    Three families, one pass: the statement and schedule rows against
    `reference.py`, the DCF/LBO forecast rows and every scalar output of the
    four valuation sheets against `valuation.py` / `comps.py` / `lbo.py`, and
    the peer table against `comps.PEERS`.
    """
    values = recalc_values(path)
    formulas = openpyxl.load_workbook(path)
    model = build_model(GREGGS_HISTORICALS, SCENARIOS[scenario])
    expected_bundle = _expected_for(scenario)
    worst = 0.0
    checked: set[str] = set()

    def check(sheet: str, col: str, row: int, label: str, expected: float) -> None:
        nonlocal worst
        checked.add(f"{sheet}!{col}{row}")
        got = values[sheet].get(f"{col}{row}")
        assert isinstance(got, (int, float)), (
            f"{scenario}: {sheet}!{col}{row} ({label}) recalculated to {got!r}"
        )
        assert got == pytest.approx(expected, abs=RECALC_TOLERANCE), (
            f"{scenario}: {sheet}!{col}{row} ({label}): {got!r} vs {expected!r}"
        )
        worst = max(worst, abs(got - expected))

    def check_within(
        sheet: str, col: str, row: int, label: str, expected: float, tolerance: float
    ) -> None:
        """Assert a cell at ITS OWN stated tolerance, outside the 1e-9 family.

        Only the peer tie rows use this. They cannot hold to 1e-9 and should
        not be expected to: each peer's market capitalisation is a transcribed
        figure rounded to £0.001m, so `price x shares - market cap` lands a few
        ten-thousandths from zero by construction. `_TIE_TOLERANCE` is the band
        `comps.py` itself ties them at, and the sheet's own label quotes it.
        Deliberately does NOT feed `worst`, which is the 1e-9 claim's number
        and would otherwise be swamped by a check held to a different standard.
        """
        checked.add(f"{sheet}!{col}{row}")
        got = values[sheet].get(f"{col}{row}")
        assert isinstance(got, (int, float)), (
            f"{scenario}: {sheet}!{col}{row} ({label}) recalculated to {got!r}"
        )
        assert abs(got - expected) < tolerance, (
            f"{scenario}: {sheet}!{col}{row} ({label}): {got!r} vs {expected!r} "
            f"(tolerance {tolerance})"
        )

    for sheet, label, expected_of in CHECKED_ROWS:
        row = _row_of(formulas[sheet], label)
        expected = expected_of(model)
        for index, col in enumerate(FORECAST_COLS):
            check(sheet, col, row, label, expected[index])

    for sheet, label, expected_of in VALUATION_ROWS:
        row = _row_of(formulas[sheet], label)
        expected = expected_of(expected_bundle)
        for index, col in enumerate(FORECAST_COLS):
            check(sheet, col, row, label, expected[index])

    for sheet, label, col, expected_of in VALUATION_CELLS:
        check(sheet, col, _row_of(formulas[sheet], label), label, expected_of(expected_bundle))

    for sheet, label, col, expected in PEER_CELLS:
        check(sheet, col, _row_of(formulas[sheet], label), label, expected)

    for sheet, label, col, expected in PEER_INPUT_CELLS + DA_INTENSITY_STAT_CELLS:
        check(sheet, col, _row_of(formulas[sheet], label), label, expected)

    for sheet, label, col, value_of in GREGGS_MARKET_CELLS:
        check(sheet, col, _row_of(formulas[sheet], label), label, value_of())

    # The peer table's two self-stated identities, for every peer.
    for label in PEER_TIE_ROWS:
        row = _row_of(formulas["Comps"], label)
        for col in PEER_COLUMNS:
            check_within("Comps", col, row, label, 0.0, TIE_TOLERANCE)

    for sheet, label, col, expected_of in LEAF_CELLS:
        check(sheet, col, _row_of(formulas[sheet], label), label, expected_of(expected_bundle))

    for label, low_of, high_of, central_of in FOOTBALL_FIELD_BARS:
        row = _row_of(formulas["Football Field"], label)
        low, high = low_of(expected_bundle), high_of(expected_bundle)
        for col, expected in (
            ("C", low),
            ("D", high),
            ("E", high - low),
            ("F", central_of(expected_bundle)),
        ):
            check("Football Field", col, row, label, expected)

    _check_sensitivity_grids(expected_bundle, check)

    return worst, checked


def _sensitivity_price(e: _Expected, wacc_rate: float, g: float) -> float:
    """The implied share price one sensitivity cell should hold.

    Composed from `valuation.py`'s own functions with the axis pair injected,
    which is the only honest way to state the expectation: re-deriving the
    terminal-year arithmetic here would test the test. Note what is and is not
    rebuilt — the terminal year IS (every intensity in it depends on ``g``), the
    explicit five years are NOT (they depend on neither axis and are merely
    re-discounted), which is exactly the split the sheet makes.
    """
    drivers = replace(e.drivers, perpetuity_growth=g)
    terminal = terminal_year(e.model, drivers, GREGGS_HISTORICALS)
    tv = terminal_value_gordon(terminal.fcf, wacc_rate, g)
    revenue = e.model.income_statement["revenue"]
    excess = e.model.balance_sheet["ppe"][-2] - terminal.ppe_intensity * revenue[-2]
    shield = excess_asset_tax_shield(
        excess, drivers.ppe_depreciation_rate, drivers.tax_rate, wacc_rate
    )
    ev = enterprise_value(unlevered_fcf(e.model, e.drivers), tv + shield, wacc_rate)
    last = GREGGS_HISTORICALS[-1]
    equity = equity_bridge(
        ev, last.borrowings.value - last.cash.value, last.lease_liabilities.value
    )
    return implied_share_price(equity, GREGGS_SHARE_COUNT.value)


def _sensitivity_multiple(e: _Expected, wacc_rate: float, g: float) -> float:
    """The terminal EV/EBITDA one cell of the second grid should hold.

    Terminal EBITDA is the LIVE one — the re-basing moves the investment and
    depreciation intensities, not the P&L margin — so only the numerator moves
    with the axes.
    """
    drivers = replace(e.drivers, perpetuity_growth=g)
    terminal = terminal_year(e.model, drivers, GREGGS_HISTORICALS)
    return terminal_value_gordon(terminal.fcf, wacc_rate, g) / e.valuation.terminal.ebitda


def _check_sensitivity_grids(e: _Expected, check) -> None:
    """All fifty grid cells and both axes, at LITERAL addresses.

    The rows and columns are hardcoded (D5:H9 and D12:H16, the growth axis on
    row 4, the WACC axis in column C) because a grid checked at addresses
    resolved from the writer would agree with a grid written one row out.
    """
    for row_index, row in enumerate(PRICE_GRID_ROWS):
        wacc_rate = e.valuation.wacc + WACC_OFFSETS[row_index]
        check("Sensitivity", "C", row, "price grid WACC axis", wacc_rate)
        for col_index, col in enumerate(GRID_COLUMNS):
            g = e.drivers.perpetuity_growth + GROWTH_OFFSETS[col_index]
            check(
                "Sensitivity", col, row, "implied share price grid",
                _sensitivity_price(e, wacc_rate, g),
            )
    for row_index, row in enumerate(MULTIPLE_GRID_ROWS):
        wacc_rate = e.valuation.wacc + WACC_OFFSETS[row_index]
        check("Sensitivity", "C", row, "multiple grid WACC axis", wacc_rate)
        for col_index, col in enumerate(GRID_COLUMNS):
            g = e.drivers.perpetuity_growth + GROWTH_OFFSETS[col_index]
            check(
                "Sensitivity", col, row, "implied terminal multiple grid",
                _sensitivity_multiple(e, wacc_rate, g),
            )
    for col_index, col in enumerate(GRID_COLUMNS):
        check(
            "Sensitivity", col, GROWTH_AXIS_ROW, "growth axis",
            e.drivers.perpetuity_growth + GROWTH_OFFSETS[col_index],
        )


@pytest.mark.parametrize("scenario", ["Bear", "Base", "Bull"])
def test_recalculated_workbook_reproduces_the_python_model(tmp_path, scenario):
    """Every cell of every statement, in every scenario, in real Excel terms."""
    path = build_workbook(GREGGS_HISTORICALS, scenario, tmp_path / f"{scenario}.xlsx")
    worst, _ = _compare_against_reference(path, scenario)
    assert worst < RECALC_TOLERANCE


# The sheets whose cells must be inside the cross-check, and why the other
# four are not. Each exclusion names where that sheet IS covered, because an
# earlier version of this comment asserted cover that did not exist — it said
# `Checks` had its own tests when no test read a single one of its results,
# and a check reading FALSE is not an error value, so the error scan did not
# see it either. Two independent reviews found it the same way.
#
# `Assumptions`: every driver is exercised by the 728-cell comparison below,
#   run in all three scenarios, so all three driver columns are read.
# `Historicals`: `test_historicals_carries_the_reported_figures_and_their_sources`,
#   plus every figure the model consumes, which arrives here through the rows
#   that ARE compared. Ten reported cells are referenced by no formula and
#   asserted by nothing — see TODO.md; they are transcription, not
#   computation, and are out of this audit's scope either way.
# `Cover`: prose. `test_the_cover_states_every_disclosure_it_must` pins the
#   content and `test_the_cover_wraps_its_notes_and_is_the_one_unfrozen_sheet`
#   the layout.
# `Checks`: TRUE/FALSE, covered by `tests/test_checks_sheet.py` — which
#   asserts every result in every scenario, that the sheet can read FALSE at
#   all, and the three disclosure figures in C13:C15.
#
# Everything else computes something a reader would quote, so everything else
# is audited here.
CALCULATION_SHEETS = (
    "IS",
    "BS",
    "CF",
    "Schedules",
    "DCF",
    "Sensitivity",
    "Comps",
    "LBO",
    "Football Field",
)

# Cells the comparison deliberately does not assert, and why each is safe.
#
# Every entry is a WORKING: a row that feeds a checked total on its own sheet,
# so a wrong value in it cannot stay hidden — it moves the total, and the total
# is compared against the Python module that owns it. Stating an expectation
# for these here would mean re-deriving the sheet's own formula in the test,
# which is the "test that re-derives a formula from its own output" this
# project has already been bitten by (see TODO.md section 8).
#
# The claim that a mutation to any of them really does surface is not asserted
# by construction — it was measured. Fourteen mutations, at least one from
# every family below, were applied to the generated workbook and recalculated:
# all fourteen failed the comparison. Anything NOT in this list and not checked
# fails the audit below, which is what stops a new row being added to a sheet
# and to no table.
CROSS_CHECK_EXEMPTIONS: tuple[tuple[str, str], ...] = (
    # Opening balances (each is the prior year's close, or FY2025 actual) and
    # the debt schedule's running subtotals. All feed a checked closing row.
    ("Schedules", "Opening net working capital"),
    ("Schedules", "Opening PP&E"),
    ("Schedules", "Opening intangibles"),
    ("Schedules", "Opening right-of-use assets"),
    ("Schedules", "Opening lease liabilities"),
    ("Schedules", "Opening shareholders' equity"),
    ("Schedules", "Opening cash"),
    ("Schedules", "Cash generated before financing"),
    ("Schedules", "Cash available before interest"),
    ("Schedules", "Cash available before debt service"),
    # The four components of terminal D&A, which sum into the checked
    # "Terminal D&A (% of revenue)", and the decay factor inside the checked
    # excess-PP&E tax shield.
    ("DCF", "Terminal PP&E depreciation (% of revenue)"),
    ("DCF", "Terminal ROU depreciation (% of revenue)"),
    ("DCF", "Terminal intangible additions (% of revenue)"),
    ("DCF", "Terminal amortisation (% of revenue)"),
    ("DCF", "Decay factor"),
    # The terminal-year workings the sensitivity grids are built from. All
    # fifty grid cells ARE checked, against `valuation.py` with the axis pair
    # injected, so a wrong working shows up as a wrong grid.
    ("Sensitivity", "Terminal PP&E / revenue"),
    ("Sensitivity", "Terminal total capex"),
    ("Sensitivity", "Terminal D&A (% of revenue: PP&E + ROU depreciation + amortisation)"),
    ("Sensitivity", "Terminal depreciation and amortisation"),
    ("Sensitivity", "Terminal EBIT "),
    ("Sensitivity", "Terminal unlevered free cash flow"),
    ("Sensitivity", "Excess PP&E over this growth rate's"),
)


def _is_computed(value) -> bool:
    return isinstance(value, (int, float)) or (
        isinstance(value, str) and value.startswith("=")
    )


def _label_matches(label: str, prefix: str) -> bool:
    """The same word-boundary rule `_row_of` uses, for the same reason.

    A bare `startswith` would exempt any future row whose label happens to
    begin with an exempted one's text — "Opening PP&E reserve added by a later
    editor" would inherit "Opening PP&E"'s exemption and never be audited.
    An exemption should cover the row it was written for and nothing else.
    """
    prefix = prefix.rstrip()
    return label == prefix or label.startswith(prefix + " ")


def test_every_computed_cell_on_a_calculation_sheet_is_inside_the_cross_check(
    workbook_path,
):
    """Coverage, proved from the cells the comparison actually asserted.

    Counting entries in the tables above would not answer this: the fifty
    sensitivity-grid cells are checked at literal addresses and appear in no
    table, and a row added to a sheet and to no table would show up in neither
    count. `_compare_against_reference` therefore records each address as it
    checks it, and this test subtracts that set from every computed cell on
    every calculation sheet. What is left must be a documented working.
    """
    _, checked = _compare_against_reference(workbook_path, "Base")
    wb = openpyxl.load_workbook(workbook_path)

    unchecked: list[str] = []
    for name in CALCULATION_SHEETS:
        ws = wb[name]
        # Column B is the label column throughout; values start at C.
        for row in ws.iter_rows(min_row=3, min_col=3):
            for cell in row:
                if not _is_computed(cell.value):
                    continue
                if f"{name}!{cell.coordinate}" in checked:
                    continue
                label = str(ws[f"B{cell.row}"].value or "").lstrip()
                if any(
                    name == sheet and _label_matches(label, prefix)
                    for sheet, prefix in CROSS_CHECK_EXEMPTIONS
                ):
                    continue
                unchecked.append(f"{name}!{cell.coordinate} ({label})")

    assert not unchecked, (
        f"{len(unchecked)} computed cell(s) are outside the cross-check and are not "
        f"a documented working. Either add them to one of the tables in this file, "
        f"or add them to CROSS_CHECK_EXEMPTIONS with the reason they are safe:\n  "
        + "\n  ".join(unchecked)
    )


def test_no_cell_anywhere_in_the_recalculated_workbook_is_an_error(workbook_path):
    """No `#REF!`, `#DIV/0!`, `Err:522` or similar, on any sheet.

    Separate from the comparison above, which only sees cells it was told to
    look at: an error in a cell no table names would otherwise reach the reader
    with nothing failing. LibreOffice surfaces these as strings in the
    recalculated file, so a scan of every cell on every sheet is the whole test.
    """
    values = recalc_values(workbook_path)

    errors = [
        f"{sheet}!{address} = {value!r}"
        for sheet, cells in values.items()
        for address, value in cells.items()
        if isinstance(value, str) and ERROR_VALUE.match(value)
    ]
    assert not errors, "error values in the recalculated workbook:\n  " + "\n  ".join(
        errors
    )


def test_toggling_only_cell_c3_re_drives_the_whole_model(workbook_path, tmp_path):
    """Change one cell, nothing else, and every driven row moves to Bull."""
    wb = openpyxl.load_workbook(workbook_path)
    assert wb["Assumptions"]["C3"].value == "Base"
    wb["Assumptions"]["C3"] = "Bull"
    toggled = tmp_path / "toggled.xlsx"
    wb.save(toggled)

    worst, _ = _compare_against_reference(toggled, "Bull")
    assert worst < RECALC_TOLERANCE

    # And it really is a different forecast, not a coincidence.
    base = build_model(GREGGS_HISTORICALS, SCENARIOS["Base"])
    bull = build_model(GREGGS_HISTORICALS, SCENARIOS["Bull"])
    assert bull.income_statement["revenue"][-1] > base.income_statement["revenue"][-1] + 300


# --------------------------------------------------------------------------
# The valuation sheets' layout, at literal addresses
# --------------------------------------------------------------------------


def test_the_valuation_sheets_put_their_first_data_row_on_row_3(workbook):
    """Same row-3 convention as the statements, on the three sheets that keep it.

    Sensitivity is the exception and is checked below: it has no year axis, so
    row 2 is blank and rows 3-4 carry the first grid's title and growth axis.
    """
    assert workbook["DCF"]["B3"].value.startswith("Cost of equity")
    assert workbook["Comps"]["B3"].value == "Share price (pence)"
    assert workbook["LBO"]["B3"].value.startswith("Entry enterprise value")


def test_comps_puts_one_peer_per_column_with_the_statistics_beside_them(workbook):
    ws = workbook["Comps"]
    assert ws["C2"].value == "SSP Group"
    assert ws["D2"].value == "Domino's Pizza Group"
    assert ws["E2"].value == "J D Wetherspoon"
    assert ws["F2"].value == "Mitchells & Butlers"
    assert ws["G2"].value == "Whitbread"
    assert (ws["I2"].value, ws["J2"].value, ws["K2"].value) == (
        "Peer min", "Peer median", "Peer max",
    )
    # The peer names on the sheet are the peer set, in its order, and no more.
    assert [p.name for p in PEERS] == [ws[f"{col}2"].value for col in PEER_COLUMNS]


def test_the_pe_statistics_exclude_ssp_by_naming_the_column_range(workbook):
    """SSP's statutory loss makes its P/E not a multiple; `multiples()` drops it.

    The sheet drops it by starting the range at Domino's rather than by a
    conditional, so the exclusion is visible in the formula bar — and this test
    reads the formula rather than the value, because a range that quietly
    included C would still produce a plausible-looking median.
    """
    ws = workbook["Comps"]
    row = _row_of(ws, "P/E")
    assert ws[f"J{row}"].value == f"=MEDIAN(D{row}:G{row})"
    # Every other statistic covers all five.
    ev_ebitda = _row_of(ws, "EV / EBITDA")
    assert ws[f"J{ev_ebitda}"].value == f"=MEDIAN(C{ev_ebitda}:G{ev_ebitda})"


def test_the_sensitivity_grids_sit_where_the_plan_puts_them(workbook):
    ws = workbook["Sensitivity"]
    assert ws["B4"].value == "Perpetuity growth (across)"
    assert ws["B5"].value == "WACC less 100bp"
    assert ws["B9"].value == "WACC plus 100bp"
    assert ws["B12"].value == "Implied terminal multiple at WACC less 100bp"
    assert ws["B16"].value == "Implied terminal multiple at WACC plus 100bp"
    # Row 2 is deliberately empty: no year axis on this sheet.
    assert all(ws[f"{col}2"].value is None for col in "BCDEFGHIJ")
    for row in list(PRICE_GRID_ROWS) + list(MULTIPLE_GRID_ROWS):
        for col in GRID_COLUMNS:
            value = ws[f"{col}{row}"].value
            assert isinstance(value, str) and value.startswith("="), f"{col}{row}"


def _copied_formula_shape(formula: str, col: str, row: int) -> str:
    """`formula` with its own column and row replaced by placeholders.

    What a formula copied across a block looks like: every reference that moves
    with the cell names the cell's own column or row, and every reference that
    does not is absolute. Substituting this cell's own coordinates therefore
    collapses a genuinely copied block to one string. The column substitution is
    bounded so it cannot touch an absolute reference to the same letter —
    ``'DCF'!$F$14`` has its F preceded by a ``$`` and must survive.
    """
    shape = re.sub(rf"(?<![$A-Z]){col}(?=\$?\d)", "<col>", formula)
    return shape.replace(f"$C{row}", "$C<row>")


def test_each_sensitivity_grid_is_one_formula_copied_across_the_block(workbook):
    """All twenty-five cells of a grid must be the same formula bar entry.

    Anchored ``$C7`` down the WACC axis and ``D$4`` across the growth axis, so
    normalising those two references should collapse the block to one string. A
    cell that reached for the wrong axis cell — the classic way to get a
    sensitivity grid subtly wrong — would survive the value comparison in one
    corner and fail here everywhere.
    """
    ws = workbook["Sensitivity"]
    for rows in (PRICE_GRID_ROWS, MULTIPLE_GRID_ROWS):
        shapes = set()
        for row in rows:
            for col in GRID_COLUMNS:
                shapes.add(_copied_formula_shape(ws[f"{col}{row}"].value, col, row))
        assert len(shapes) == 1, f"grid rows {rows} hold {len(shapes)} different formulas"


def test_the_exit_multiple_row_carries_its_caveat_on_the_sheet(workbook):
    """The warning has to be IN the workbook, not only in the source.

    A reader who takes the exit-multiple terminal value for a second opinion is
    being misled — it is the peer median EV/EBIT times terminal EBIT and nothing
    else — so the note sits on the row below it, and the two memo rows that
    follow are the arithmetic proof.
    """
    labels = [c.value for c in workbook["DCF"]["B"] if isinstance(c.value, str)]
    caveat = next(label for label in labels if label.startswith("READ THIS"))
    for phrase in (
        "NOT an independent second method",
        "identity",
        "peer median EV/EBIT x terminal EBIT",
        "builds no part of the enterprise value",
    ):
        assert phrase in caveat, f"the DCF caveat does not say {phrase!r}"
    assert labels.index(caveat) > labels.index("Terminal value (exit multiple)")
