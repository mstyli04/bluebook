from pathlib import Path

import openpyxl
import pytest

from bluebook.inputs.greggs import GREGGS_HISTORICALS
from bluebook.workbook.build import build_workbook

BAR_LABELS = (
    "DCF (Gordon) — WACC +/-100bp, perpetuity growth +/-50bp",
    "Trading comps — peer minimum to peer maximum EV/EBITDA",
    "52-week traded range — the market's own answer",
)


@pytest.fixture(scope="module")
def workbook(tmp_path_factory) -> openpyxl.Workbook:
    path = build_workbook(
        GREGGS_HISTORICALS, "Base", tmp_path_factory.mktemp("ff") / "model.xlsx"
    )
    return openpyxl.load_workbook(path)


def _bar_rows(ws) -> list[int]:
    wanted = set(BAR_LABELS)
    return [c.row for c in ws["B"] if c.value in wanted]


def test_the_sheet_carries_all_three_bars(workbook):
    ws = workbook["Football Field"]
    labels = {c.value for c in ws["B"] if isinstance(c.value, str)}
    for required in BAR_LABELS:
        assert required in labels, f"football field missing bar '{required}'"


def test_every_bar_cell_is_a_formula(workbook):
    """`Football Field` is not in HARDCODE_ALLOWED — nothing here may be a constant."""
    ws = workbook["Football Field"]
    for row in _bar_rows(ws):
        for col in ("C", "D", "E", "F"):
            value = ws[f"{col}{row}"].value
            assert isinstance(value, str) and value.startswith("="), (
                f"{col}{row} is not a formula: {value!r}"
            )


def test_the_span_is_high_less_low_on_its_own_row(workbook):
    """The chart's visible series reads the span; it must not drift off its row."""
    ws = workbook["Football Field"]
    for row in _bar_rows(ws):
        assert ws[f"E{row}"].value == f"=D{row}-C{row}"


def test_there_is_no_exit_multiple_bar(workbook):
    """See Task 13 item 5: it is one peer statistic, not a second method."""
    ws = workbook["Football Field"]
    labels = [c.value for c in ws["B"] if isinstance(c.value, str)]
    bars = [label for label in labels if label in BAR_LABELS]
    assert len(bars) == 3
    assert not any(
        "exit multiple" in label.lower() and label in BAR_LABELS for label in labels
    )
    # ...and the omission is explained on the sheet rather than left silent.
    assert any(
        isinstance(c.value, str) and "NO exit-multiple bar" in c.value
        for c in ws["B"]
    )


def test_the_staleness_of_the_traded_range_is_disclosed(workbook):
    ws = workbook["Football Field"]
    assert any(
        isinstance(c.value, str) and "wrong tomorrow" in c.value for c in ws["B"]
    )


def test_the_chart_exists_and_is_a_stacked_horizontal_bar(workbook):
    ws = workbook["Football Field"]
    assert len(ws._charts) == 1, "expected exactly one chart on the football field"
    chart = ws._charts[0]
    assert chart.type == "bar"
    assert chart.grouping == "stacked"
    assert chart.overlap == 100
    assert len(chart.series) == 2, "a floating bar needs a pedestal and a span series"
    assert chart.series[0].graphicalProperties.noFill is True


def test_the_dcf_bar_spans_the_whole_sensitivity_grid(workbook):
    """Low and high must read all five grid rows, not just the centre one."""
    ws = workbook["Football Field"]
    row = _bar_rows(ws)[0]
    for col, function in (("C", "MIN"), ("D", "MAX")):
        formula = ws[f"{col}{row}"].value
        assert formula.startswith(f"={function}(")
        assert formula.count("'Sensitivity'!") == 5, (
            f"{col}{row} should reference all five grid rows: {formula}"
        )
