from pathlib import Path

import openpyxl

from bluebook.inputs.greggs import GREGGS_HISTORICALS
from bluebook.workbook.build import build_workbook


def test_build_produces_all_expected_sheets(tmp_path: Path):
    path = build_workbook(GREGGS_HISTORICALS, "Base", tmp_path / "model.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [
        "Cover", "Checks", "Assumptions", "Historicals",
        "IS", "BS", "CF", "Schedules",
        "DCF", "Sensitivity", "Comps", "LBO", "Football Field",
    ]


def test_forecast_cells_are_formulas_not_constants(tmp_path: Path):
    path = build_workbook(GREGGS_HISTORICALS, "Base", tmp_path / "model.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["IS"]
    for row in ws.iter_rows(min_col=6, max_col=10):
        for cell in row:
            if cell.value is not None:
                assert isinstance(cell.value, str) and cell.value.startswith("="), (
                    f"IS!{cell.coordinate} holds a constant: {cell.value!r}"
                )
