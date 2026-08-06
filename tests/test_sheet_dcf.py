from pathlib import Path

import openpyxl
import pytest

from bluebook.inputs.greggs import GREGGS_HISTORICALS
from bluebook.workbook.build import build_workbook
from bluebook.workbook.layout import Layout


def test_dcf_sheet_has_the_key_output_rows(tmp_path: Path):
    path = build_workbook(GREGGS_HISTORICALS, "Base", tmp_path / "model.xlsx")
    wb = openpyxl.load_workbook(path)
    labels = {
        c.value for c in wb["DCF"]["B"] if isinstance(c.value, str)
    }
    for required in (
        "WACC", "Unlevered free cash flow", "Terminal value (Gordon growth)",
        "Terminal value (exit multiple)", "Enterprise value",
        "Equity value", "Implied share price (p)",
    ):
        assert required in labels, f"DCF sheet missing '{required}'"


def test_sensitivity_grid_is_fully_populated_with_formulas(tmp_path: Path):
    path = build_workbook(GREGGS_HISTORICALS, "Base", tmp_path / "model.xlsx")
    ws = openpyxl.load_workbook(path)["Sensitivity"]
    grid = [
        c for row in ws.iter_rows(min_row=5, max_row=9, min_col=4, max_col=8)
        for c in row
    ]
    assert len(grid) == 25
    assert all(isinstance(c.value, str) and c.value.startswith("=") for c in grid)
