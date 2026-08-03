import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import pytest

from bluebook.recalc import RecalcError, recalc, recalc_values


@pytest.fixture
def simple_workbook(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 2.0
    ws["A2"] = 3.0
    ws["A3"] = "=A1*A2"
    ws["A4"] = "=SUM(A1:A2)"
    path = tmp_path / "simple.xlsx"
    wb.save(path)
    return path


def test_openpyxl_writes_no_cached_value(simple_workbook: Path):
    """Baseline: the generated file has formulas but no computed results."""
    wb = openpyxl.load_workbook(simple_workbook, data_only=True)
    assert wb["Sheet1"]["A3"].value is None


def test_recalc_values_computes_formulas(simple_workbook: Path):
    values = recalc_values(simple_workbook)
    assert values["Sheet1"]["A3"] == pytest.approx(6.0)
    assert values["Sheet1"]["A4"] == pytest.approx(5.0)


def test_recalc_returns_persisted_recalculated_file(simple_workbook: Path):
    """recalc() itself (not just recalc_values()) must produce a readable,
    persisted xlsx carrying LibreOffice's computed cache. recalc() does not
    clean this up -- that's the caller's job, exercised here directly."""
    recalculated = recalc(simple_workbook)
    try:
        assert recalculated.exists()
        assert recalculated.name == simple_workbook.name
        wb = openpyxl.load_workbook(recalculated, data_only=True)
        assert wb["Sheet1"]["A3"].value == pytest.approx(6.0)
        wb.close()
    finally:
        shutil.rmtree(recalculated.parent, ignore_errors=True)


def test_recalc_values_cleans_up_its_temp_directory(simple_workbook: Path):
    """recalc_values() must not leak the outer mkdtemp() directory that
    recalc() created for it -- only the profile subdirectory was being
    removed before this fix, leaving the outer dir (and the copied xlsx
    inside it) behind on every call."""
    tmp_root = Path(tempfile.gettempdir())
    before = set(tmp_root.glob("bluebook-recalc-*"))

    recalc_values(simple_workbook)

    after = set(tmp_root.glob("bluebook-recalc-*"))
    leaked = after - before
    assert not leaked, f"recalc_values() leaked temp directories: {leaked}"


def test_recalc_values_handles_multiple_worksheets(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1["A1"] = 10.0
    ws2 = wb.create_sheet("Second")
    ws2["A1"] = 20.0
    ws2["B1"] = "=A1*2"
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    values = recalc_values(path)

    assert set(values.keys()) == {"First", "Second"}
    assert values["First"]["A1"] == pytest.approx(10.0)
    assert values["Second"]["A1"] == pytest.approx(20.0)
    assert values["Second"]["B1"] == pytest.approx(40.0)


def test_recalc_values_preserves_string_and_bool_types(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["A2"] = 5.0
    ws["A3"] = 5.0
    ws["A4"] = "=A2=A3"  # boolean formula result
    path = tmp_path / "types.xlsx"
    wb.save(path)

    values = recalc_values(path)

    assert values["Sheet1"]["A1"] == "hello"
    assert isinstance(values["Sheet1"]["A1"], str)
    assert values["Sheet1"]["A4"] is True
    assert isinstance(values["Sheet1"]["A4"], bool)


def test_recalc_raises_recalc_error_when_soffice_executable_missing(
    simple_workbook: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(
        "bluebook.recalc.SOFFICE", "definitely-not-a-real-soffice-binary"
    )
    with pytest.raises(RecalcError, match="definitely-not-a-real-soffice-binary"):
        recalc(simple_workbook)


def test_recalc_raises_recalc_error_on_timeout(
    simple_workbook: Path, monkeypatch: pytest.MonkeyPatch
):
    def fake_run(*args, **kwargs):
        raise subprocess.TimeoutExpired(cmd="soffice", timeout=1, output="", stderr="")

    monkeypatch.setattr("bluebook.recalc.subprocess.run", fake_run)
    with pytest.raises(RecalcError, match="timed out"):
        recalc(simple_workbook)


def test_recalc_raises_recalc_error_on_nonzero_returncode(
    simple_workbook: Path, monkeypatch: pytest.MonkeyPatch
):
    def fake_run(*args, **kwargs):
        return subprocess.CompletedProcess(
            args=args, returncode=1, stdout="", stderr="boom"
        )

    monkeypatch.setattr("bluebook.recalc.subprocess.run", fake_run)
    with pytest.raises(RecalcError, match="exit code 1"):
        recalc(simple_workbook)
